from utils.utils import find_files
import pandas as pd
import os
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from data_transform.add_pivots import add_pivots
from openpyxl import load_workbook
import re
from data_loader.master_data_reader import get_data_from_rt
from utils.helper import is_processed_today, mark_processed_today
from config import RT_REFERENCE_PATH

# --- Utility untuk tulis DF ke sheet ---
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

def _pydatefmt_to_excel(fmt: str) -> str:
    """
    Konversi sederhana Python date format (e.g. "%Y-%m-%d %H:%M:%S")
    ke Excel date format (e.g. "yyyy-mm-dd hh:mm:ss").

    Hanya mengganti token umum. Jika string tidak mengandung '%', 
    diasumsikan sudah format Excel (tetap dikembalikan apa adanya).
    """
    if not isinstance(fmt, str):
        return "yyyy-mm-dd"
    if "%" not in fmt:
        return fmt  # diasumsikan sudah excel-like

    # mapping token python -> excel
    replacements = [
        ("%Y", "yyyy"),
        ("%y", "yy"),
        ("%m", "mm"),
        ("%d", "dd"),
        ("%H", "hh"),
        ("%I", "hh"),
        ("%M", "mm"),   # minutes; Excel uses mm but disambiguated by presence of hh
        ("%S", "ss"),
        ("%p", "AM/PM"),
        # timezone/other tokens ignored
    ]

    out = fmt
    for pytok, ex in replacements:
        out = out.replace(pytok, ex)
    # Replace double spaces, ensure clean separators
    out = re.sub(r"\s+", " ", out).strip()
    return out or "yyyy-mm-dd"

def _write_df_to_sheet(ws, df_data: pd.DataFrame, sheet_title: str, date_cols, excel_date_format="dd/mm/yyyy"):
        try:
            ws.title = sheet_title[:31]  # max 31 char
        except Exception:
            ws.title = "Sheet"
        for r in dataframe_to_rows(df_data, index=False, header=True):
            ws.append(r)
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                try:
                    cell.border = thin_border
                except Exception:
                    pass
                if cell.row == 1:
                    cell.fill = header_fill
                else:
                    # formatting tanggal
                    col_header = headers[cell.column - 1] if (0 <= cell.column - 1 < len(headers)) else None
                    if col_header in date_cols and isinstance(cell.value, datetime):
                        cell.number_format = excel_date_format
        # auto adjust col width
        for col_cells in ws.columns:
            try:
                col_letter = get_column_letter(col_cells[0].column)
            except Exception:
                continue
            max_length = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > 200:
                    val = val[:200]
                max_length = max(max_length, len(val))
            ws.column_dimensions[col_letter].width = max_length + 2

def _sanitize_sheet_name(name: str) -> str:
    """Bersihkan nama sheet dari karakter ilegal & limit panjang"""
    if not name or str(name).strip() == "":
        return "Unknown"

    # Hilangkan karakter ilegal Excel
    cleaned = re.sub(r"[\[\]\*:/\\\?']", "", str(name))

    # Hapus leading/trailing spasi
    cleaned = cleaned.strip()

    # Default jika kosong setelah bersih
    if not cleaned:
        cleaned = "Unknown"

    # Batasi 31 karakter
    return cleaned[:31]

def _safe_save_wb(wb, output_path, max_retries=5, delay=1):
    """
    Coba simpan workbook, jika gagal PermissionError (file sedang dipakai),
    fallback simpan ke nama baru dengan suffix (1), (2), dst.
    """
    attempt = 0
    base, ext = os.path.splitext(output_path)
    final_path = output_path

    while attempt < max_retries:
        try:
            wb.save(final_path)
            return final_path
        except PermissionError:
            attempt += 1
            final_path = f"{base} ({attempt}){ext}"
            print(f"⚠️ {output_path} sedang dipakai, coba simpan sebagai {final_path}")
            time.sleep(delay)

    raise PermissionError(f"Gagal menyimpan setelah {max_retries} percobaan: {output_path}")

def save_with_styling_and_date(df: pd.DataFrame, output_path: str, edit_file: dict, manuals_path: str = None, debug: bool = True):
    """
    Save DataFrame ke Excel dengan styling:
    - replace manuals.csv overrides
    - konversi date/numeric
    - tulis Master + optional split
    - validasi hasil simpan
    """

    if debug:
        print(f"[DEBUG] save_with_styling_and_date START -> {output_path}")

    df = df.copy()

    # --- Konversi kolom tanggal ---
    date_cols = []
    for col in df.columns:
        if "tgl" in col.lower() or "date" in col.lower() or "eta" in col.lower() or "tanggal" in col.lower():
            try:
                if debug: print(f"[DEBUG] parsing date column: {col}")
                df[col] = pd.to_datetime(df[col], errors="coerce")
                date_cols.append(col)
            except Exception as e:
                print(f"⚠️ Kolom {col} gagal diformat: {e}")

    # --- Konversi kolom numerik ---
    for col in ["QTY", "WEIGHT", "AMOUNT", "BILNOTE_AMOUNT"]:
        if col in df.columns:
            try:
                df[col] = pd.to_numeric(df[col], errors="coerce")
            except Exception:
                if debug: print(f"[DEBUG] numeric conv failed for {col}, left as-is")

    # --- Setup Workbook ---
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)  # remove default

    # --- Tentukan format tanggal Excel ---
    raw_date_format = edit_file.get("date_format")
    excel_date_format = _pydatefmt_to_excel(raw_date_format) if raw_date_format else "dd/mm/yyyy"
    if debug:
        print(f"[DEBUG] using excel_date_format: {excel_date_format!r}")

    # --- Tambah Master sheet SELALU (pivot butuh ini) ---
    ws_master = wb.create_sheet(title=edit_file.get("rename_data_sheet", "Master"))
    _write_df_to_sheet(ws_master, df, ws_master.title,
                       date_cols=date_cols, excel_date_format=excel_date_format)

    # --- Split kalau ada config ---
    split_cfg = edit_file.get("split_sheet_by_column")
    if split_cfg and isinstance(split_cfg, dict):
        split_col = split_cfg.get("column")

        if debug:
            print(f"[DEBUG] Splitting sheet by column: {split_col}")

        if split_col in df.columns:
            col_lower = split_col.lower()

            if "tgl" in col_lower or "date" in col_lower:
                if debug: print(f"[DEBUG] Kolom {split_col} dianggap tanggal, split by MONTH")
                df["_split_month"] = df[split_col].dt.strftime("%B").str.upper()
                df["_split_month_num"] = df[split_col].dt.month
                group_col = "_split_month"
                unique_vals = (
                    df.dropna(subset=["_split_month_num"])
                      .sort_values("_split_month_num")["_split_month"]
                      .unique()
                )
            else:
                group_col = split_col
                unique_vals = df[group_col].dropna().unique()

            for val in unique_vals:
                df_group = df[df[group_col] == val]
                sheet_name = _sanitize_sheet_name(str(val) if pd.notna(val) else "Unknown")
                ws_group = wb.create_sheet(title=sheet_name)
                _write_df_to_sheet(
                    ws_group,
                    df_group.drop(columns=["_split_month", "_split_month_num"], errors="ignore"),
                    ws_group.title,
                    date_cols=date_cols,
                    excel_date_format=excel_date_format,
                )
        else:
            print(f"⚠️ split_sheet_by_column gagal, kolom {split_col} tidak ditemukan!")

    # --- Split by specific column values ---
    split_val_cfg = edit_file.get("split_sheet_by_col_val")

    if split_val_cfg:
        # normalize ke list
        if isinstance(split_val_cfg, dict):
            split_val_cfg = [split_val_cfg]

        for cfg in split_val_cfg:
            col = cfg.get("col")
            values = cfg.get("val", [])

            # --- find actual column (case-insensitive) ---
            matched_col = None
            if isinstance(col, str):
                for c in df.columns:
                    try:
                        if str(c).strip().lower() == col.strip().lower():
                            matched_col = c
                            break
                    except Exception:
                        continue
            else:
                matched_col = col

            if matched_col is None:
                print(f"⚠️ split_sheet_by_col_val gagal, kolom {col} tidak ditemukan!")
                continue

            # --- prepare comparison series ---
            series = df[matched_col].astype(object).where(
                pd.notna(df[matched_col]), other=pd.NA
            )
            series_str = series.astype(str).str.strip()

            def _make_mask(target):
                if target is None or (isinstance(target, str) and target.strip() == ""):
                    return (
                        series.isna()
                        | (series_str == "nan")
                        | (series_str == "None")
                        | (series_str == "")
                    )
                return series_str == str(target).strip()

            # --- dict mode: {value: sheet_name} ---
            if isinstance(values, dict):
                for v, sheet_alias in values.items():
                    mask = _make_mask(v)
                    df_group = df[mask]

                    if not df_group.empty:
                        sheet_name = _sanitize_sheet_name(str(sheet_alias))
                        ws_group = wb.create_sheet(title=sheet_name)

                        _write_df_to_sheet(
                            ws_group,
                            df_group,
                            ws_group.title,
                            date_cols=date_cols,
                            excel_date_format=excel_date_format,
                        )

                        print(
                            f"📝 Sheet '{sheet_name}' dibuat "
                            f"dengan filter {matched_col} = {v}"
                        )

            # --- list mode: [value1, value2, ...] ---
            else:
                for v in values:
                    mask = _make_mask(v)
                    df_group = df[mask]

                    if not df_group.empty:
                        sheet_name = _sanitize_sheet_name(str(v))
                        ws_group = wb.create_sheet(title=sheet_name)

                        _write_df_to_sheet(
                            ws_group,
                            df_group,
                            ws_group.title,
                            date_cols=date_cols,
                            excel_date_format=excel_date_format,
                        )

                        print(
                            f"📝 Sheet '{sheet_name}' dibuat "
                            f"dengan filter {matched_col} = {v}"
                        )

    # --- Tambah sheet AWB RT kalau ada config ---
    awb_rt_cfg = edit_file.get("awb_rt_sheet")
    if awb_rt_cfg and isinstance(awb_rt_cfg, dict):
        rt_cols = awb_rt_cfg.get("col", [])
        try:
            rt_df = get_data_from_rt(output_path, rt_cols, RT_REFERENCE_PATH, debug=debug)
            if not rt_df.empty:
                ws_rt = wb.create_sheet(title="AWB RT")
                _write_df_to_sheet(ws_rt, rt_df, ws_rt.title,
                                date_cols=date_cols,
                                excel_date_format=excel_date_format)
                print(f"📝 Sheet 'AWB RT' berhasil ditambahkan")
        except Exception as e:
            print(f"⚠️ Gagal menambahkan sheet AWB RT: {e}")

    # --- Save & validate (pakai safe save) ---
    try:
        final_path = _safe_save_wb(wb, output_path, max_retries=3)
        wb.close()
    except Exception as e:
        print(f"❌ Gagal menyimpan file {output_path}: {e}")
        raise

    try:
        _ = load_workbook(filename=final_path, read_only=True, keep_links=False)
    except Exception as e:
        print(f"❌ VALIDATION ERROR: {e}")
        raise

    main_sheet_name = edit_file.get("rename_data_sheet", "Master")
    print(f"✅ Disimpan dengan style + edit → {final_path} (sheet: {main_sheet_name})")
    return main_sheet_name, final_path

def process_edit_file(criteria_lists, output_dir: str, manuals_path: str = "./data/manuals.csv", bypass_history: bool = False):
    os.makedirs(output_dir, exist_ok=True)
    frontline_files = {}

    for criteria in criteria_lists:
        saved_as = criteria.get("save_as", criteria.get("group_name"))
        edit_file = criteria.get("edit_file", {})
        pic_frontline = criteria.get("pic_frontline")

        print(f"\n🔍 Mencari file untuk group: {saved_as}")
        files_found = find_files(output_dir, saved_as)
        if not files_found:
            print(f"❌ Tidak ada file ditemukan untuk {saved_as}")
            continue

        # Jika ada file duplikat seperti "TADA.xlsx" dan "TADA (1).xlsx",
        # proses hanya file terbaru agar file stale tidak menimpa hasil terbaru.
        latest_by_key = {}
        for p in files_found:
            try:
                folder = os.path.dirname(p)
                stem = os.path.splitext(os.path.basename(p))[0]
                canonical_stem = re.sub(r"\s*\(\d+\)$", "", stem).strip().upper()
                key = (folder, canonical_stem)
                mtime = os.path.getmtime(p)
                prev = latest_by_key.get(key)
                if prev is None or mtime > prev[0]:
                    latest_by_key[key] = (mtime, p)
            except Exception:
                # fallback: tetap masukkan jika metadata gagal dibaca
                key = (os.path.dirname(p), os.path.basename(p).upper())
                latest_by_key[key] = (0, p)

        files_found = [v[1] for v in latest_by_key.values()]
        files_found.sort()
        if len(files_found) > 1:
            print(f"ℹ️ Ditemukan beberapa varian file {saved_as}, proses file terbaru per nama dasar.")

        for file_path in files_found:
            # Buat process_key unik per file
            filename = os.path.basename(file_path)
            process_key = f"edit_{filename}"
            actual_file_path = file_path

            # === Cek apakah file ini sudah pernah diproses hari ini ===
            if not bypass_history and is_processed_today(process_key):
                print(f"⏩ Skip edit {file_path}, sudah diproses hari ini.")
                if pic_frontline:
                    if isinstance(pic_frontline, (list, tuple, set)):
                        for person in pic_frontline:
                            frontline_files.setdefault(person, []).append(actual_file_path)
                    else:
                        frontline_files.setdefault(pic_frontline, []).append(actual_file_path)
                continue

            print(f"✅ File ditemukan: {file_path}")

            # === Proses edit file ===
            df = None
            try:
                # Try standard read with openpyxl
                df = pd.read_excel(file_path, dtype=str, engine="openpyxl")
            except Exception as e1:
                # If failed, try with data_only=True (read cached values only, skip formulas)
                try:
                    print(f"   ⚠️ Retrying with data_only=True...")
                    wb = load_workbook(file_path, data_only=True)
                    ws = wb.active
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        data.append(row)
                    if data:
                        df = pd.DataFrame(data[1:], columns=data[0])
                        df = df.astype(str)
                except Exception as e2:
                    print(f"⚠️ Gagal membaca file {file_path}: {e1}")
                    print(f"   Retry juga gagal: {e2}")
                    print(f"   ➡️ File mungkin corrupted, skip file ini.")
                    continue
            
            if df is None or df.empty:
                print(f"⚠️ File kosong atau tidak bisa dibaca: {file_path}")
                continue

            if edit_file:
                sheet_name, actual_file_path = save_with_styling_and_date(
                    df, file_path, edit_file, manuals_path, debug=False
                )

                # jeda kecil biar stabil
                time.sleep(1.5)

                # --- Tambah pivot ---
                if edit_file.get("pivots"):
                    ok = add_pivots(
                        file_path=actual_file_path,
                        data_sheet_name=sheet_name,
                        pivots=edit_file["pivots"],
                        summary_sheet_name="Summary",
                        date_format=edit_file.get("date_format", "dd/mm/yyyy"),
                    )
                    if ok:
                        print(f"📊 PivotTables ditambahkan ke {actual_file_path}")
                    else:
                        print(f"⚠️ Pivot gagal untuk {actual_file_path}")

                # --- Hapus Master kalau tidak diminta ---
                split_cfg = edit_file.get("split_sheet_by_column")
                if split_cfg and not split_cfg.get("include_master", False):
                    try:
                        wb = load_workbook(actual_file_path)
                        if "Master" in wb.sheetnames:
                            wb.remove(wb["Master"])
                            wb.save(actual_file_path)
                        wb.close()
                    except Exception as e:
                        print(f"⚠️ Gagal menghapus Master dari {actual_file_path}: {e}")

            if pic_frontline:
                if isinstance(pic_frontline, (list, tuple, set)):
                    for person in pic_frontline:
                        frontline_files.setdefault(person, []).append(actual_file_path)
                else:
                    frontline_files.setdefault(pic_frontline, []).append(actual_file_path)

            # Tandai sudah diproses hari ini
            mark_processed_today(process_key)
            if actual_file_path != file_path:
                mark_processed_today(f"edit_{os.path.basename(actual_file_path)}")

    return frontline_files

def is_file_unlocked(path, timeout=30):
    start = time.time()
    while time.time() - start < timeout:
        try:
            with open(path, "r+b"):
                return True
        except PermissionError:
            time.sleep(1)
        except FileNotFoundError:
            return False
    return False