from datetime import datetime
from dateutil.relativedelta import relativedelta
import os
import pandas as pd
import calendar
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from data_transform.add_columns import add_address, add_status_pod, add_cust_name, add_coding_delivery, add_eta, add_grouping_late, add_origin_city, add_periode, add_province_zipcode, add_status_latlong, add_date_time_received, add_grouping_sla, add_PIC, add_SBS, add_dept_PZC, add_AJCar_status, add_is_close, add_reason, add_no, add_reason_last_attempt, add_date_receive_request, add_descr_return, add_update_time, add_uob_pickup_data_cols, add_sociolla_cols, add_grouping_status, fix_regional_cols, add_aging_carrer
from data_transform.parse_dates import normalize_all_dates
from .ref_data_loader import load_cust_ref
from utils.helper import save_history, load_history, mark_processed_today, is_processed_today
from utils.utils import clean_receiver_column
from config import TABLE_REFERENCE_PATH

# ------------------- KONSTANTA -------------------
INDO_MONTHS = [
    "", "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
    "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"
]

EN_TO_ID_MONTH = {
    "JANUARY": "JANUARI", "FEBRUARY": "FEBUARI", "MARCH": "MARET",
    "APRIL": "APRIL", "MAY": "MEI", "JUNE": "JUNI", "JULY": "JULI",
    "AUGUST": "AGUSTUS", "SEPTEMBER": "SEPTEMBER", "OCTOBER": "OKTOBER",
    "NOVEMBER": "NOVEMBER", "DECEMBER": "DESEMBER"
}

# ------------------- MAPPING KOMPONEN TRANSFORMASI -------------------
TRANSFORM_GROUPS = {
    # Group yang hasilkan banyak kolom sekaligus
    "ORIGIN_CITY_GROUP": {
        "cols": ["ORIGIN_CITY", "ORIGIN_CITY_2", "3_LC_ORIGIN", "PROVINSI_ORIGIN", "REGION", "KODE_POS_ORI"],
        "func": lambda df, ref: add_origin_city(df, ref),
    },
    "PROVINCE_GROUP": {
        "cols": ["KECAMATAN", "NAMA KAB/KOTA", "PROVINSI", "KODE_POS"],
        "func": lambda df, ref: add_province_zipcode(df, ref),
    },
    "PERIODE_GROUP": {
        "cols": ["DAY", "PERIODE", "WEEK", "WEEK_OF_YEAR"],
        "func": lambda df, ref: add_periode(df),
    },
    "TGL_RECEIVED_GROUP": {
        "cols": ["DATE_RECEIVED", "TIME_RECEIVED"],
        "func": lambda df, ref: add_date_time_received(df),
    },
    "AGING_CARRER_GROUP": {
        "cols": ["AGING_1ST", "CAREER_1ST", "AGING_POD", "CARRER_POD"],
        "func": lambda df, ref: add_aging_carrer(df),
    },
    "AJ_CAR_GROUP": {
        "cols": ["AJ Car Status", "KETERANGAN AJ CAR", "Remark Return"],
        "func": lambda df, ref: add_AJCar_status(df, ref),
    },
    "UOB_GROUP": {
        "cols": ["CYCLE NAME", "REFF NUM", "PICK UP DATE", "CUSTOMER NAME", "RECIPIENT RELATION", "STATUS UOB", "REASON 2"],
        "func": lambda df, ref: add_uob_pickup_data_cols(df)
    },
    "SOCIOLLA_GROUP": {
        "cols": ["DIM", "DIM - Copy"],
        "func": lambda df, ref: add_sociolla_cols(df)
    }
}

# Fungsi yang hanya hasilkan satu kolom
TRANSFORM_FUNCS = {
    "ADDRESS": lambda df, ref: add_address(df, "ADDRESS"),
    "STATUS_POD": lambda df, ref: add_status_pod(df),
    "CUST_NAME": lambda df, ref: add_cust_name(df),
    "STATUS_LATLONG": lambda df, ref: add_status_latlong(df),
    "ETA": lambda df, ref: add_eta(df),
    "CODING_DELIVERY": lambda df, ref: add_coding_delivery(df, ref),
    "GROUPING_LATE": lambda df, ref: add_grouping_late(df),
    "GROUPING_SLA": lambda df, ref: add_grouping_sla(df),
    "PIC": lambda df, ref: add_PIC(df),
    "ORIGIN 2": lambda df, ref: add_SBS(df),
    "DEPT PZC": lambda df, ref: add_dept_PZC(df),
    "IS_CLOSE": lambda df, ref: add_is_close(df),
    "REASON": lambda df, ref: add_reason(df),
    "REASON_LAST_ATTEMPT": lambda df, ref: add_reason_last_attempt(df, ref),
    "DATE_RECEIVE_REQUEST": lambda df, ref: add_date_receive_request(df),
    "DESCR_RETURN": lambda df, ref: add_descr_return(df),
    "UPDATE_TIME": lambda df, ref: add_update_time(df),
    "GROUPING_STATUS": lambda df, ref: add_grouping_status(df)
}

# daftar kolom yang kamu butuhkan
REQUIRED_COLS = [
    "AWB","ID_ACCOUNT","SHIPPER_NAME","TGL_ENTRY","CONSIGNEE_NAME","ADDR1","ADDR2","ADDR3",
    "CONTACT","NOTELP","NOREF","ORIGIN","DEST","SERVICE","QTY","WEIGHT","GOODS_DESCR",
    "INSURANCE_ID","GOODS_VALUE","INSURANCE_VALUE(+)","AMOUNT","INTRUCTION","NOTICE",
    "HOLD_REASON","RECEIVING","RECEIVING_DATE","OUTBOUND_MANIFEST","OUTBOUND_MANIFEST_DATE",
    "INBOUND_MANIFEST","USER_IM","INBOUND_MANIFEST_DATE","MANIFEST_TRANSIT_AGEN","DATE_TRANSIT",
    "HVO_NO","HVO_DATE","HVO_HUB","HVO_HUB_NAME","HVO_HUB_DESTINATION","HVO_HUB_DESTINATION_NAME",
    "HVI_NO","HVI_DATE","RUNSHEET_NO","DATE_RUNSHEET","RUNSHEET_COURIER_ID","RUNSHEET_COURIER_NAME",
    "CODING","STATUS_POD","TGL_RECEIVED","STATUS_LATITUDE","STATUS_LONGITUDE","AGING","ETD","SLA",
    "CARRER","RECEIVED/REASON","TGL_UPDATE_STATUS_POD","WUS_OUTGOING_CODE","WUS_REMARKS","WUS_DATE",
    "INVOICED","AWB_CANCEL","COD_FLAG","BILNOTE_FLAG","BILNOTE_AMOUNT","REFNO_UOB","SCO_NO",
    "WO/DO/PO","NO_INVOICE","PAYMENT_TYPE","DATE_1ST_ATTEMPT","RESULT_1ST_ATTEMPT","LATLONG_1ST_ATTEMPT",
    "DATE_2ND_ATTEMPT","RESULT_2ND_ATTEMPT","LATLONG_2ND_ATTEMPT","DATE_LAST_ATTEMPT","RESULT_LAST_ATTEMPT",
    "LATLONG_LAST_ATTEMPT","PRA_RUNSHEET_NO","PRA_RUNSHEET_NAME","PRA_RUNSHEET_DATE","CS3_DATE",
    "CONNOTE_RETURN_RT","DATE_CONNOTE_RETURN_RT","CONNOTE_RETURN_RF","DATE_CONNOTE_RETURN_RF","USER_CONNOTE",
    "USER_ZONE_CONNOTE","CONFIRM_SHIPMENT_UNDEL","TRANSIT_MANIFEST","TRANSIT_MANIFEST_DATE","TRANSIT_MANIFEST_USER",
    "IREG_MANIFEST","IREG_CODE","IREG_DATE","URL_TTD","URL_FOTO","USER_OM","USER_RECEIVING","AGING_ONGOING",
    "CLAIM_NO","CLAIM_DOC_NO","CLAIM_DATE","NO_CNOTE_FW","ORIGIN_FW","DEST_FW","CODING_STATUS_FW","DESC_STATUS_FW",
    "HBG_NO","HBG_DATE","1ST_HVO_NO","1ST_HVO_DATE","1ST_HVO_USER","LAST_HVO_NO","LAST_HVO_DATE","LAST_HVO_USER",
    "MANIFEST_TRANSIT_SUBAGEN_NO","MANIFEST_TRANSIT_SUBAGEN_DATE","MANIFEST_INBOUND_SUBAGEN_NO","MANIFEST_INBOUND_SUBAGEN_DATE",
    "BAG_NO","LATEST_SM_NO","LATEST_SM_DATE","1ST_PREVIOUS_SM_NO","1ST_PREVIOUS_SM_DATE","2ND_PREVIOUS_SM_NO","2ND_PREVIOUS_SM_DATE",
    "1ST_TRANSIT_MANIFEST_NO","1ST_TRANSIT_MANIFEST_DATE","2ND_TRANSIT_MANIFEST_NO","2ND_TRANSIT_MANIFEST_DATE",
    "3RD_TRANSIT_MANIFEST_NO","3RD_TRANSIT_MANIFEST_DATE","LAST_TRANSIT_MANIFEST_NO","LAST_TRANSIT_MANIFEST_DATE",
    "MTI_USER","MTS_USER","HO_COURIER_NO","HO_COURIER_DATE","WAREHOUSE_DATE","OFFICE_DATE","IRREG_REMAKS","BPIK","ZONE_USER_ENTRI",
    "CORRECT_DESTINATION","CORRECT_SERVICE","CORRECT_AMOUNT","HACB_NO","HACB_DATE","HACB_USER","HBAG_NO","HBAG_DATE","HBAG_USER",
    "PICKUP_DATE","PICKUP_STATUS","PICKUP_COURIER_ID","1ST_RUNSHEET_DATE","1ST_RUNSHEET_COURIERID","URL_CHAT","SINGLE_LEG",
    "LAST_DATE_DO","NO_RCW","DATE_RCW","USER_RCW","DATE_LPR","NO_LPR","NO_RDO","DATE_RDO","NO_DO","PROJECT_KR","HO_OFFICE_NO",
    "HO_OFFICE_DATE","LATEST_SM_ORIGIN","LATEST_SM_DEST","1ST_PREVIOUS_SM_ORIGIN","1ST_PREVIOUS_SM_DEST","2ND_PREVIOUS_SM_ORIGIN",
    "2ND_PREVIOUS_SM_DEST","TGL_TARIK_REPORT","RESPONCIBILITY","STATUS_VERSI_CCC","STATUS_POD_UPDATE","1ST_ATTEMPT_DATE","AGING_1ST",
    "CAREER_1ST","AGING_POD","CARRER_POD","CODING_UNDEL","REASON RETURN","3 LC DEST","NAMA KAB/KOTA 2","REGIONAL","ZONA",
    "GROUPING_SHIPPER","CATEGORY","REFERENCE CUST CCC","PAYMENT_METHODE","CUST_INDUSTRY","BIG_GROUPING_CUST","PIC_NAME_NEW RELATION",
    "PIC SUPPORT DATA","UNIT","DEPT","DATE","PERIODE","PERIODE_WEEK","ORIGIN RT","DEST RT","3LC Last Status","Regional Last Status",
    "Zona Last Status","OTS by CT","PERIODE_OTS","CLOSING_OTS","CODING_RT","RECEIVED/REASON_RT","Return Date"
]

table_reference_path = TABLE_REFERENCE_PATH
# ------------------- SAVE DENGAN SHORT DATE -------------------
def save_with_shortdate(df: pd.DataFrame, output_path="output.xlsx"):
    df = df.copy()
    df = df.where(pd.notna(df), None)
    
    # kolom yang tidak boleh di-parse jadi datetime
    exclude_cols = ["status_pod_update"]

    # konversi numeric columns agar tidak berubah string
    NUMERIC_COLS = ["QTY", "WEIGHT", "AMOUNT", "BILNOTE_AMOUNT"]
    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    
    # pastikan kolom tanggal memang datetime (defensive)
    for col in [c for c in df.columns if any(k in c.lower() for k in ["tgl", "date"]) and c.lower() not in exclude_cols]:
        try:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        except Exception:
            pass

    if "NO" in df.columns:
        df = add_no(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    thin_border = Border(
        left=Side(style="thin"), 
        right=Side(style="thin"),
        top=Side(style="thin"), 
        bottom=Side(style="thin")
    )

    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # loop semua sel untuk border + format tanggal
    for row in ws.iter_rows():
        for cell in row:
            # Tambah border
            cell.border = thin_border
            # Tambah warna header
            if cell.row == 1:
                cell.fill = header_fill
            # Format tanggal
            if isinstance(cell.value, (datetime, pd.Timestamp)):
                cell.number_format = "MM/DD/YYYY"

    # auto-adjust column width
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_length = max(len(str(cell.value) if cell.value else "") for cell in col_cells)
        ws.column_dimensions[col_letter].width = max_length + 2

    # --- safe save dengan fallback ---
    output_path = Path(output_path)
    file_stem, file_ext = output_path.stem, output_path.suffix
    file_dir = output_path.parent

    attempt = 0
    while True:
        try:
            candidate = file_dir / f"{file_stem}{'' if attempt == 0 else f' ({attempt})'}{file_ext}"
            wb.save(candidate)
            _update_status_history(df, candidate)
            break
        except PermissionError:
            attempt += 1
            continue

def style_worksheet(ws):
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
            if isinstance(cell.value, (datetime, pd.Timestamp)):
                cell.number_format = "MM/DD/YYYY"

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_length = max(len(str(cell.value) if cell.value else "") for cell in col_cells)
        ws.column_dimensions[col_letter].width = max_length + 2

def _update_status_history(df: pd.DataFrame, saved_path: str):
    """Update status_pod history setelah file disimpan"""
    if "STATUS_POD" not in df.columns:
        return
    
    today_counts = df["STATUS_POD"].value_counts().to_dict()
    basename = os.path.basename(saved_path)
    history = load_history()
    history[basename] = today_counts
    save_history(history)

def insert_insurance_id(cols):
        cols = list(cols)

        if "INSURANCE_ID" in cols:
            return cols

        # Jika GOODS_DESCR ada → sisip setelahnya
        if "GOODS_DESCR" in cols:
            idx = cols.index("GOODS_DESCR") + 1
            return cols[:idx] + ["INSURANCE_ID"] + cols[idx:]

        # Cari kolom standar terakhir yang muncul di selected_cols
        existing_std = [c for c in REQUIRED_COLS if c in cols]
        if existing_std:
            last_std = existing_std[-1]
            idx = cols.index(last_std) + 1
            return cols[:idx] + ["INSURANCE_ID"] + cols[idx:]

        # Jika tidak ada kolom standard → sisip di depan
        return ["INSURANCE_ID"] + cols

# ------------------- FUNGSI MAIN -------------------
def get_recent_month_paths(base_dir, bulan_ke_belakang=3, category=str):
    today = datetime.today()
    folders = []
    for i in range(bulan_ke_belakang):
        target_date = today - relativedelta(months=i)
        tahun = target_date.year
        bulan = target_date.month
        nama_bulan = INDO_MONTHS[bulan]
        nama_folder_bulan = f"{bulan}. {nama_bulan} {tahun}"
        full_path = os.path.join(base_dir, str(tahun), nama_folder_bulan, "CATEGORY", category)
        folders.append(full_path)
    return folders

def get_data_from_master(base_dir, criteria=None, category=str, output_dir=None, reference_path=str, bypass_history=False, debug=False):
    combined_df = pd.DataFrame()
    base_unc = r"\\?\UNC" + base_dir[1:] if base_dir.startswith(r"\\") else base_dir
    # --- Ambil bulan_awal dari criteria (jika ada) ---
    bulan_awal = None
    if isinstance(criteria, dict):
        bulan_awal = criteria.get("bulan_awal")
    elif isinstance(criteria, list) and len(criteria) > 0:
        bulan_awal = criteria[0].get("bulan_awal")

    # --- Hitung bulan_ke_belakang berdasarkan bulan_awal ---
    bulan_ke_belakang = 3  # default
    if bulan_awal:
        bulan_awal = bulan_awal.strip().lower()
        bulan_map = {
            **{m.lower(): i for i, m in enumerate(calendar.month_name) if m},  # English
            "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5,
            "juni": 6, "juli": 7, "agustus": 8, "september": 9,
            "oktober": 10, "november": 11, "desember": 12
        }

        if bulan_awal in bulan_map:
            start_month = bulan_map[bulan_awal]
            current_month = datetime.today().month
            bulan_ke_belakang = (current_month - start_month) + 1
            if bulan_ke_belakang <= 0:
                bulan_ke_belakang += 12  # handle kalau awal tahun ke akhir tahun
            print(f"🗓️ Mengambil data dari bulan {bulan_awal.title()} hingga bulan ini ({bulan_ke_belakang} bulan).")
        else:
            print(f"⚠️ Nama bulan tidak dikenali: {bulan_awal}, pakai default 3 bulan terakhir.")

    # --- Dapatkan folder dengan fungsi existing ---
    target_folders = get_recent_month_paths(base_unc, bulan_ke_belakang=bulan_ke_belakang, category=category)

    for folder in target_folders:
        if not os.path.exists(folder):
            print(f"📂 Folder tidak ditemukan: {folder}")
            continue

        print(f"📂 Membaca file {category} di: {folder}")

        all_dfs = []
        for fname in os.listdir(folder):
            if fname.lower().endswith(".csv"):
                file_path = os.path.join(folder, fname)
                try:
                    df = pd.read_csv(file_path, encoding="utf-8", low_memory=False)
                    df.columns = df.columns.str.strip().str.upper()
                    all_dfs.append(df)
                except Exception as e:
                    print(f"⚠️ Gagal baca {fname}: {e}")

        if not all_dfs:
            continue

        folder_df = pd.concat(all_dfs, ignore_index=True, join="outer")
        folder_df = folder_df[[c for c in REQUIRED_COLS if c in folder_df.columns]]
        combined_df = pd.concat([combined_df, folder_df], ignore_index=True, join="outer")

    if combined_df.empty:
        print("❌ Tidak ada data berhasil digabung.")
        return
    
    # --- Deduplicate AWB by latest TGL_TARIK_REPORT ---
    if "AWB" in combined_df.columns:
        before = len(combined_df)
        if "TGL_TARIK_REPORT" in combined_df.columns:
            combined_df["TGL_TARIK_REPORT"] = pd.to_datetime(
                combined_df["TGL_TARIK_REPORT"], errors="coerce"
            )
            combined_df = combined_df.sort_values(
                by=["AWB", "TGL_TARIK_REPORT"], ascending=[True, True]
            )
        combined_df = combined_df.drop_duplicates(subset=["AWB"], keep="last")
        after = len(combined_df)
        if before != after:
            print(f"🧹 Duplikat dihapus: {before - after} baris (tinggal {after})")

    print(f"📊 Total data gabungan: {len(combined_df)} baris")

    combined_df = normalize_all_dates(combined_df, debug=False)
    combined_df = fix_regional_cols(combined_df)

    if not criteria:
        return combined_df

    if isinstance(criteria, dict):
        criteria = [criteria]

    save_base = output_dir or base_unc
    os.makedirs(save_base, exist_ok=True)

    for crit in criteria:
        try:
            group_name = crit.get("group_name")
            id_account = crit.get("id_account")
            selected_cols = crit.get("selected_cols")
            exclude_cols = crit.get("exclude_cols")
            split_by_id = crit.get("split_by_id", False)
            split_by_col_val = crit.get("split_by_col_val")
            save_as = crit.get("save_as")
            jumlah_bulan = crit.get("jumlah_bulan")
            selected_statuses = crit.get("selected_statuses")
            filter_cols = crit.get("filter_cols")
            move_to_file = crit.get("move_to_sheet_of_file")
            clean_receiver = crit.get("clean_receiver")
            rename_cols = crit.get("rename_cols")
            saved_as = crit.get("save_as") or group_name
            cust_ref = crit.get("cust_ref") or {}
            selected_ref_cols = cust_ref.get("selected_ref_cols", [])
            ref_sheet = cust_ref.get("ref_sheet", [])
            date_col = cust_ref.get("date_col", "TGL_ENTRY")
            process_key = f"query_{saved_as}"

            if not bypass_history and is_processed_today(process_key):
                print(f"⏩ Skip {saved_as}, sudah diproses hari ini (gunakan bypass_history=True untuk paksa ulang).")
                continue

            print(f"\n🔍 Memfilter untuk grup: {group_name}")
            if isinstance(group_name, list):
                df_filtered = combined_df[combined_df["BIG_GROUPING_CUST"].isin(group_name)].copy()
            else:
                df_filtered = combined_df[combined_df["BIG_GROUPING_CUST"] == group_name].copy()

            if id_account and "ID_ACCOUNT" in df_filtered.columns:
                df_filtered = df_filtered[df_filtered["ID_ACCOUNT"].isin(id_account)]

            if selected_statuses and "STATUS_POD" in df_filtered.columns:
                df_filtered = df_filtered[df_filtered["STATUS_POD"].isin(selected_statuses)]

            if jumlah_bulan:
                awal_bulan_ini = datetime.today().replace(day=1)
                bulan_awal = awal_bulan_ini - relativedelta(months=jumlah_bulan - 1)
                bulan_akhir = awal_bulan_ini + relativedelta(months=1)
                df_filtered = df_filtered[
                    (df_filtered["TGL_ENTRY"] >= bulan_awal) & 
                    (df_filtered["TGL_ENTRY"] < bulan_akhir)
                ]

            if clean_receiver:
                df_filtered["RECEIVED/REASON"] = df_filtered.apply(clean_receiver_column, axis=1)

            # transformasi kolom
            if selected_cols:
                # insert "INSURANCE_ID" to selected_cols
                selected_cols = insert_insurance_id(selected_cols)

                # pastikan df_filtered punya kolom INSURANCE_ID
                if "INSURANCE_ID" not in df_filtered.columns:
                    df_filtered["INSURANCE_ID"] = ""

                needed_cols = set(selected_cols)
                for tg_name, group in TRANSFORM_GROUPS.items():
                    if any(col in needed_cols for col in group["cols"]):
                        if df_filtered.empty:
                            continue
                        try:
                            df_filtered = group["func"](df_filtered, reference_path)
                        except Exception as e:
                            print(f"⚠️ Gagal menambahkan {tg_name}: {e}")

                for col in needed_cols:
                    if col in TRANSFORM_FUNCS:
                        if df_filtered.empty:
                            continue
                        try:
                            df_filtered = TRANSFORM_FUNCS[col](df_filtered, reference_path)
                        except Exception as e:
                            print(f"⚠️ Gagal menambahkan {col}: {e}")

                for col in selected_cols:
                    if col not in df_filtered.columns:
                        df_filtered[col] = ""

            if filter_cols:
                for col_name, ftype, arg in filter_cols:
                    if col_name not in df_filtered.columns:
                        continue
                    if ftype == "starts_with":
                        df_filtered = df_filtered[df_filtered[col_name].astype(str).str.startswith(str(arg))]
                    elif ftype == "ends_with":
                        df_filtered = df_filtered[df_filtered[col_name].astype(str).str.endswith(str(arg))]
                    elif ftype == "contains":
                        df_filtered = df_filtered[df_filtered[col_name].astype(str).str.contains(str(arg), case=False, na=False)]
                    elif ftype == "equals":
                        df_filtered = df_filtered[df_filtered[col_name] == arg]
                    elif ftype == "not_equals":
                        df_filtered = df_filtered[df_filtered[col_name] != arg]
                    elif ftype == "gt":
                        df_filtered = df_filtered[df_filtered[col_name].astype(float) > float(arg)]
                    elif ftype == "lt":
                        df_filtered = df_filtered[df_filtered[col_name].astype(float) < float(arg)]
                    elif ftype == "in":
                        df_filtered = df_filtered[df_filtered[col_name].isin(arg)]
                    elif ftype == "not_in":
                        df_filtered = df_filtered[~df_filtered[col_name].isin(arg)]

            if cust_ref:
                ref_path = cust_ref.get("ref_path")
                key_left, key_right = cust_ref.get("key_col", ["AWB", "AWB"])
                jumlah_bulan = cust_ref.get("jumlah_bulan", 3)

                print(f"🔄 Membaca customer reference dari: {ref_path}")
                df_custref = load_cust_ref(ref_path, jumlah_bulan, ref_sheet)
                df_custref = normalize_all_dates(df_custref, debug=False)

                # NORMALISASI KEY
                for df_, key in [(df_custref, key_left), (df_filtered, key_right)]:
                    if key in df_.columns:
                        df_[key] = (
                            df_[key]
                            .astype(str)
                            .str.strip()
                            .str.replace(r"\s+", "", regex=True)
                            .str.upper()
                            .apply(lambda x: x if x.startswith("'") else f"'{x}")
                        )

                if not df_custref.empty:
                    df_merged = df_custref.merge(
                        df_filtered,
                        how="left",
                        left_on=key_left,
                        right_on=key_right,
                        suffixes=("", "_M")
                    )

                    # drop kolom _M
                    df_merged = df_merged[[c for c in df_merged.columns if not c.endswith("_M")]]

                    df_filtered = df_merged

                    if "NO" in df_filtered.columns:
                        df_filtered = add_no(df_filtered)

                    # ============================================
                    #  Tambah SELECTED_COLS_REF ke SELECTED_COLS master
                    # ============================================
                    if selected_ref_cols:
                        # gabungkan list ref + list master
                        combined_cols = list(dict.fromkeys(selected_ref_cols + selected_cols))

                        selected_cols = combined_cols


            if df_filtered.empty:
                print(f"⚠️ Tidak ada data cocok untuk grup: {group_name}")
                continue

            # -----------------------
            # helper lokal: apply exclude + select + rename tepat sebelum save
            # -----------------------
            def prepare_df_for_save(df_segment: pd.DataFrame) -> pd.DataFrame:
                df_out = df_segment.copy()
                # drop kolom bantuan / sensitive dulu kalau diminta
                if exclude_cols:
                    df_out.drop(columns=exclude_cols, errors="ignore", inplace=True)
                # pilih kolom sesuai selected_cols (jika ada)
                if selected_cols:
                    cols_to_keep = [col for col in selected_cols if col in df_out.columns]
                    df_out = df_out[cols_to_keep]
                    # rename jika rename_cols spesifik dan panjang sesuai
                    if rename_cols:
                        if len(rename_cols) != len(selected_cols):
                            # hanya rename kolom yang ada di selected_cols yang tersisa
                            # buat mapping minimal (zip only on common length) — tapi beri warning
                            print(f"⚠️ Panjang rename_cols ≠ selected_cols untuk grup {group_name}, skip full rename")
                        else:
                            rename_map = {old: new for old, new in zip(selected_cols, rename_cols)}
                            # hanya apply pada kolom yang ada di df_out
                            rename_map = {k: v for k, v in rename_map.items() if k in df_out.columns}
                            if rename_map:
                                df_out.rename(columns=rename_map, inplace=True)
                                print(f"🔄 Kolom di-rename untuk grup {group_name}")
                else:
                    # jika tidak ada selected_cols tapi ada rename_cols (jarang), coba rename intersect
                    if rename_cols:
                        # can't reliably map without selected_cols => skip but warn
                        print(f"⚠️ rename_cols disediakan tanpa selected_cols untuk grup {group_name} → skip rename")
                return df_out

            # === Save handling ===
            edit_file = crit.get("edit_file", {})

            # --- PRIORITAS: split_by_id ---
            if split_by_id and "ID_ACCOUNT" in df_filtered.columns:
                for id_acc, df_id in df_filtered.groupby("ID_ACCOUNT"):
                    clean_id_acc = str(id_acc).lstrip("'")

                    if jumlah_bulan and date_col in df_id.columns:
                        df_id["BULAN_STR"] = df_id["TGL_ENTRY"].dt.strftime("%B").str.upper()
                        for bulan_eng, df_bln in df_id.groupby("BULAN_STR"):
                            nama_bulan = EN_TO_ID_MONTH.get(bulan_eng, bulan_eng.title())
                            fname = f"{save_as or group_name} - {clean_id_acc} - {nama_bulan}.xlsx"
                            output_path = os.path.join(save_base, fname)

                            # prepare then save
                            df_to_save = prepare_df_for_save(df_bln.drop(columns=["BULAN_STR"], errors="ignore"))
                            if edit_file:
                                df_to_save.to_excel(output_path, index=False)
                                _update_status_history(df_bln, output_path)
                            else:
                                save_with_shortdate(df_to_save, output_path)
                            print(f"✅ Disimpan: {output_path}")
                            mark_processed_today(process_key)
                    else:
                        fname = f"{save_as or group_name} - {clean_id_acc}.xlsx"
                        output_path = os.path.join(save_base, fname)

                        df_to_save = prepare_df_for_save(df_id)
                        if edit_file:
                            df_to_save.to_excel(output_path, index=False)
                            _update_status_history(df_id, output_path)
                        else:
                            save_with_shortdate(df_to_save, output_path)
                        print(f"✅ Disimpan: {output_path}")
                        mark_processed_today(process_key)
                continue

            # --- PRIORITAS KEDUA: split_by_col_val ---
            if split_by_col_val and split_by_col_val in df_filtered.columns:
                for val, df_val in df_filtered.groupby(split_by_col_val):
                    clean_val = str(val).replace("/", "-").replace("\\", "-").strip()

                    if jumlah_bulan and date_col in df_val.columns:
                        df_val["BULAN_STR"] = df_val["TGL_ENTRY"].dt.strftime("%B").str.upper()
                        for bulan_eng, df_bln in df_val.groupby("BULAN_STR"):
                            nama_bulan = EN_TO_ID_MONTH.get(bulan_eng, bulan_eng.title())
                            fname = f"{save_as or group_name} - {clean_val} - {nama_bulan}.xlsx"
                            output_path = os.path.join(save_base, fname)

                            df_to_save = prepare_df_for_save(df_bln.drop(columns=["BULAN_STR"], errors="ignore"))
                            if edit_file:
                                df_to_save.to_excel(output_path, index=False)
                                _update_status_history(df_bln, output_path)
                            else:
                                save_with_shortdate(df_to_save, output_path)
                            print(f"✅ Disimpan: {output_path}")
                            mark_processed_today(process_key)
                    else:
                        fname = f"{save_as or group_name} - {clean_val}.xlsx"
                        output_path = os.path.join(save_base, fname)

                        df_to_save = prepare_df_for_save(df_val)
                        if edit_file:
                            df_to_save.to_excel(output_path, index=False)
                            _update_status_history(df_val, output_path)
                        else:
                            save_with_shortdate(df_to_save, output_path)
                        print(f"✅ Disimpan: {output_path}")
                        mark_processed_today(process_key)
                continue

            # --- PRIORITAS KETIGA: move_to_file (save ke sheet) ---
            if move_to_file:
                target_file = os.path.join(save_base, f"{move_to_file}.xlsx")
                sheet_name = save_as or group_name

                df_to_save = prepare_df_for_save(df_filtered)

                if not os.path.exists(target_file):
                    # create new file with sheet
                    save_with_shortdate(df_to_save, target_file, sheet_name=sheet_name)
                    print(f"✅ Dibuat file baru {target_file} dengan sheet {sheet_name}")
                    mark_processed_today(process_key)
                else:
                    book = openpyxl.load_workbook(target_file)
                    if sheet_name in book.sheetnames:
                        book.remove(book[sheet_name])
                    ws = book.create_sheet(title=sheet_name)

                    # tulis header & rows dari df_to_save
                    for i, col in enumerate(df_to_save.columns, 1):
                        ws.cell(row=1, column=i, value=col)
                    for r_idx, row in enumerate(df_to_save.itertuples(index=False), 2):
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)

                    style_worksheet(ws)
                    book.save(target_file)
                    print(f"📄 Data disimpan ke {target_file} (sheet: {sheet_name})")
                    mark_processed_today(process_key)
                continue

            # --- fallback: pecah per bulan atau single file ---
            if jumlah_bulan and date_col in df_filtered.columns:
                df_filtered["BULAN_STR"] = df_filtered[date_col].dt.strftime("%B").str.upper()
                for bulan_eng, df_bln in df_filtered.groupby("BULAN_STR"):
                    nama_bulan = EN_TO_ID_MONTH.get(bulan_eng, bulan_eng.title())
                    fname = f"{save_as or group_name} - {nama_bulan}.xlsx"
                    output_path = os.path.join(save_base, fname)

                    df_to_save = prepare_df_for_save(df_bln.drop(columns=["BULAN_STR"], errors="ignore"))
                    if edit_file:
                        df_to_save.to_excel(output_path, index=False)
                        _update_status_history(df_bln, output_path)
                    else:
                        save_with_shortdate(df_to_save, output_path)
                    print(f"✅ Disimpan: {output_path}")
                    mark_processed_today(process_key)
            else:
                fname = f"{save_as or group_name}.xlsx"
                output_path = os.path.join(save_base, fname)

                df_to_save = prepare_df_for_save(df_filtered)
                if edit_file:
                    df_to_save.to_excel(output_path, index=False)
                    _update_status_history(df_filtered, output_path)
                else:
                    save_with_shortdate(df_to_save, output_path)
                print(f"✅ Disimpan: {output_path}")
                mark_processed_today(process_key)

        except Exception as e:
            print(f"❌ Gagal proses {crit.get('group_name')}: {e}")
            continue

def get_data_from_rt(master_path: str, selected_cols: list, rt_master_path: str,
                     reference_path=table_reference_path, debug: bool=False) -> pd.DataFrame:
    # --- Load Master ---
    if debug: print(f"[DEBUG] Load Master dari {master_path}")
    try:
        df_master = pd.read_excel(master_path, dtype=str)
    except Exception as e:
        print(f"⚠️ Gagal membaca master: {e}")
        return pd.DataFrame(columns=selected_cols)

    if "CONNOTE_RETURN_RT" not in df_master.columns:
        print("⚠️ Kolom CONNOTE_RETURN_RT tidak ada di Master!")
        return pd.DataFrame(columns=selected_cols)
    
    df_master = df_master[["CONNOTE_RETURN_RT"]].copy()

    # --- Load RT Master ---
    if debug: print(f"[DEBUG] Load RT Master dari {rt_master_path}")
    try:
        df_rt = pd.read_excel(rt_master_path, dtype=str)
    except Exception as e:
        print(f"⚠️ Gagal membaca RT Master: {e}")
        return pd.DataFrame(columns=selected_cols)

    df_rt.columns = df_rt.columns.str.strip().str.upper()
    join_key = "AWB"
    if "AWB - COPY" in df_rt.columns:
        join_key = "AWB - COPY"

    # --- Join Master (keep semua kolom RT Master) ---
    df_joined = df_master.merge(
        df_rt,
        left_on="CONNOTE_RETURN_RT",
        right_on=join_key,
        how="inner"
    )

    if df_joined.empty:
        print("⚠️ Tidak ada AWB yang match di RT Master.")
        return pd.DataFrame(columns=selected_cols)

    # --- Load Reference Table ---
    if debug: print(f"[DEBUG] Load Reference Table dari {reference_path}")
    try:
        df_ref = pd.read_excel(reference_path, dtype=str)
        df_ref.columns = df_ref.columns.str.strip().str.upper()
    except Exception as e:
        print(f"⚠️ Gagal membaca reference: {e}")
        df_ref = pd.DataFrame()

    df_out = df_joined.copy()

    # --- Apply TRANSFORM_GROUPS ---
    for tg_name, group in TRANSFORM_GROUPS.items():
        if any(col in selected_cols for col in group["cols"]):
            if debug: print(f"Menambahkan kolom group {tg_name} ({group['cols']})..")
            try:
                df_out = group["func"](df_out, df_ref)
            except Exception as e:
                print(f"⚠️ Gagal menambahkan {tg_name}: {e}")

    # --- Apply TRANSFORM_FUNCS ---
    for col in selected_cols:
        if col in TRANSFORM_FUNCS:
            if debug: print(f"Menambahkan kolom {col}..")
            try:
                df_out = TRANSFORM_FUNCS[col](df_out, df_ref)
            except Exception as e:
                print(f"⚠️ Gagal menambahkan {col}: {e}")

    # --- Tambahkan kolom kosong kalau masih hilang ---
    for col in selected_cols:
        if col not in df_out.columns:
            df_out[col] = ""

    # --- Urutkan sesuai selected_cols ---
    df_out = df_out[[col for col in selected_cols if col in df_out.columns]]

    return df_out

