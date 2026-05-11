from datetime import datetime
from dateutil.relativedelta import relativedelta
import os
import re
import time
import pandas as pd
import numpy as np
import duckdb
import openpyxl
from openpyxl.styles import Border, Side, PatternFill
from data_transform.add_columns import (
    add_address, add_status_pod_c01, add_cust_name, add_coding_delivery, add_eta,
    add_grouping_late, add_origin_city, add_periode, add_province_zipcode, add_reason_undel,
    add_status_latlong, add_date_time_received, add_grouping_sla, add_PIC,
    add_SBS, add_dept_PZC, add_AJCar_status, add_is_close, add_reason, add_no,
    add_reason_last_attempt, add_date_receive_request, add_descr_return, add_rodamas_cols,
    add_update_time, add_uob_pickup_data_cols, add_sociolla_cols, add_grouping_status, add_aging_carrer,
    add_SPK, add_reason_1st_attempt, add_wilayah, add_young_living_cols, fix_contact_notelp_col, fix_empty_date_1st_attempt, add_3lc_dest_fw, add_rounded_weight, add_bni_kategori, add_coding_remarks
)
from data_transform.parse_dates import normalize_all_dates
from .ref_data_loader import load_cust_ref
from utils.helper import save_history, load_history, mark_processed_today, is_processed_today
from utils.utils import clean_receiver_column
from config import TABLE_REFERENCE_PATH

# ------------------- KONSTANTA -------------------
INDO_MONTHS = [
    "", "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
    "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"
]

EN_TO_ID_MONTH = {
    "JANUARY": "JANUARI", "FEBRUARY": "FEBRUARI", "MARCH": "MARET",
    "APRIL": "APRIL", "MAY": "MEI", "JUNE": "JUNI", "JULY": "JULI",
    "AUGUST": "AGUSTUS", "SEPTEMBER": "SEPTEMBER", "OCTOBER": "OKTOBER",
    "NOVEMBER": "NOVEMBER", "DECEMBER": "DESEMBER"
}

# ------------------- MAPPING KOMPONEN TRANSFORMASI -------------------
# (preserve your mappings)
TRANSFORM_GROUPS = {
    "ORIGIN_CITY_GROUP": {
        "cols": ["ORIGIN_CITY", "ORIGIN_CITY_2", "3_LC_ORIGIN", "PROVINSI_ORIGIN", "REGION", "KODE_POS_ORI"],
        "func": lambda df, ref: add_origin_city(df, ref),
    },
    "PROVINCE_GROUP": {
        "cols": ["KECAMATAN", "NAMA KAB/KOTA", "PROVINSI", "KODE POS"],
        "func": lambda df, ref: add_province_zipcode(df, ref),
    },
    "PERIODE_GROUP": {
        "cols": ["DAY", "PERIODE", "WEEK", "WEEK_OF_YEAR", "TIME_ENTRY"],
        "func": lambda df, ref: add_periode(df),
    },
    "TGL_RECEIVED_GROUP": {
        "cols": ["DATE_RECEIVED", "TIME_RECEIVED"],
        "func": lambda df, ref: add_date_time_received(df),
    },
    "AGING_CARRER_GROUP": {
        "cols": ["AGING_1ST", "CAREER_1ST", "AGING_POD", "CARRER_POD"],
        "func": lambda df, ref: add_aging_carrer(df),
    },
    "AJ_CAR_GROUP": {
        "cols": ["AJ Car Status", "KETERANGAN AJ CAR", "Remark Return"],
        "func": lambda df, ref: add_AJCar_status(df, ref),
    },
    "UOB_GROUP": {
        "cols": ["CYCLE NAME", "REFF NUM", "PICK UP DATE", "CUSTOMER NAME", "RECIPIENT RELATION", "STATUS UOB", "REASON 2"],
        "func": lambda df, ref: add_uob_pickup_data_cols(df)
    },
    "SOCIOLLA_GROUP": {
        "cols": ["DIM", "DIM - Copy"],
        "func": lambda df, ref: add_sociolla_cols(df)
    },
    "YOUNG_LIVING_GROUP": {
        "cols": ["OTMS Order", "Target", "Origin_WH", "Vendor", "No_YL", "No DSV", "Buyer PO", "Indonesia Bagian", "City Code", "Harga Per KG", "ATA", "LEAD TIME", "SLA_", "REMARK_YL", "AWB 2"],
        "func": lambda df, ref: add_young_living_cols(df)
    },
    "RODAMAS_GROUP": {
        "cols": ["NAMA ORIGIN", "CUSTOMER"],
        "func": lambda df, ref: add_rodamas_cols(df)
    },
    "FIX_CONTACT_NOTELP_GROUP": {
        "cols": ["CONTACT", "NOTELP"],
        "func": lambda df, ref: fix_contact_notelp_col(df)
    }
}

TRANSFORM_FUNCS = {
    "ADDRESS": lambda df, ref: add_address(df, "ADDRESS"),
    "STATUS_POD": lambda df, ref: add_status_pod_c01(df),
    "CUST_NAME": lambda df, ref: add_cust_name(df),
    "STATUS_LATLONG": lambda df, ref: add_status_latlong(df),
    "ETA": lambda df, ref: add_eta(df),
    "CODING_DELIVERY": lambda df, ref: add_coding_delivery(df, ref),
    "GROUPING_LATE": lambda df, ref: add_grouping_late(df),
    "GROUPING_SLA": lambda df, ref: add_grouping_sla(df),
    "PIC": lambda df, ref: add_PIC(df),
    "ORIGIN 2": lambda df, ref: add_SBS(df),
    "DEPT PZC": lambda df, ref: add_dept_PZC(df),
    "IS_CLOSE": lambda df, ref: add_is_close(df),
    "REASON": lambda df, ref: add_reason(df),
    "REASON_LAST_ATTEMPT": lambda df, ref: add_reason_last_attempt(df, ref),
    "DATE_RECEIVE_REQUEST": lambda df, ref: add_date_receive_request(df),
    "DESCR_RETURN": lambda df, ref: add_descr_return(df),
    "UPDATE_TIME": lambda df, ref: add_update_time(df),
    "GROUPING_STATUS": lambda df, ref: add_grouping_status(df),
    "SPK": lambda df, ref: add_SPK(df),
    "REASON_1ST_ATTEMPT": lambda df, ref: add_reason_1st_attempt(df, ref),
    "WILAYAH": lambda df, ref: add_wilayah(df, ref),
    "DATE_1ST_ATTEMPT": lambda df, ref: fix_empty_date_1st_attempt(df),
    "3 LC DEST FW": lambda df, ref: add_3lc_dest_fw(df),
    "WEIGHT": lambda df, ref: add_rounded_weight(df),
    "REASON UNDEL": lambda df, ref: add_reason_undel(df, ref),
    "BNI_KATEGORI": lambda df, ref: add_bni_kategori(df),
    "CODING_REMARKS": lambda df, ref: add_coding_remarks(df)
}

STATUS_POD_FULL_INPUT_COLS = [
    "CODING", "AWB_CANCEL", "TGL_ENTRY", "CONNOTE_RETURN_RT", "CONNOTE_RETURN_RF",
    "RESULT_1ST_ATTEMPT", "STATUS_POD", "STATUS_POD_UPDATE", "INBOUND_MANIFEST",
    "MANIFEST_TRANSIT_AGEN", "HVI_NO", "RUNSHEET_NO", "OUTBOUND_MANIFEST",
    "RECEIVING", "HVO_NO", "ORIGIN", "DEST", "CONFIRM_SHIPMENT_UNDEL",
    "DATE_RUNSHEET", "NO_CNOTE_FW", "DEST_FW", "CODING_STATUS_FW",
    "DESC_STATUS_FW", "IREG_CODE",
]

TRANSFORM_INPUT_COLS = {
    "ADDRESS": ["ADDR1", "ADDR2", "ADDR3"],
    "STATUS_POD": ["CODING", "STATUS_POD", "STATUS_POD_UPDATE"],
    "CUST_NAME": ["GROUPING_SHIPPER"],
    "STATUS_LATLONG": ["STATUS_LATITUDE", "STATUS_LONGITUDE"],
    "ETA": ["ETD", "TGL_ENTRY"],
    "CODING_DELIVERY": ["CODING"],
    "GROUPING_LATE": [
        "ORIGIN", "3 LC DEST", "OUTBOUND_MANIFEST_DATE", "1ST_HVO_DATE",
        "HBG_DATE", "PICKUP_STATUS", "PICKUP_DATE", "TGL_ENTRY",
        "INBOUND_MANIFEST_DATE", "HVI_DATE", "1ST_RUNSHEET_DATE", "ETD",
        "ZONA", "MANIFEST_TRANSIT_SUBAGEN_DATE", "MANIFEST_INBOUND_SUBAGEN_DATE",
        "RECEIVING_DATE", "TGL_RECEIVED", "DATE_LAST_ATTEMPT", "STATUS_POD",
        "NO_CNOTE_FW", "HOLD_REASON", "CARRER",
    ],
    "GROUPING_SLA": ["SLA"],
    "ORIGIN 2": ["ORIGIN"],
    "DEPT PZC": ["ID_ACCOUNT", "GROUPING_SHIPPER"],
    "IS_CLOSE": ["STATUS_POD"],
    "REASON": ["STATUS_POD", "AGING_POD"],
    "REASON_LAST_ATTEMPT": ["RESULT_LAST_ATTEMPT"],
    "DATE_RECEIVE_REQUEST": ["CONNOTE_RETURN_RT", "TGL_RECEIVED", "DATE_CONNOTE_RETURN_RT"],
    "DESCR_RETURN": ["STATUS_POD"],
    "GROUPING_STATUS": ["STATUS_POD"],
    "SPK": ["INTRUCTION"],
    "REASON_1ST_ATTEMPT": ["RESULT_1ST_ATTEMPT"],
    "DATE_1ST_ATTEMPT": ["DATE_1ST_ATTEMPT", "STATUS_POD", "TGL_RECEIVED"],
    "3 LC DEST FW": ["DEST_FW"],
    "REASON UNDEL": ["CODING_UNDEL"],
    "BNI_KATEGORI": ["GOODS_DESCR"],
    "CODING_REMARKS": ["CODING"]
}

TRANSFORM_GROUP_INPUT_COLS = {
    "ORIGIN_CITY_GROUP": ["ORIGIN"],
    "PROVINCE_GROUP": ["DEST"],
    "PERIODE_GROUP": ["TGL_ENTRY"],
    "TGL_RECEIVED_GROUP": ["TGL_RECEIVED"],
    "AJ_CAR_GROUP": STATUS_POD_FULL_INPUT_COLS,
    "UOB_GROUP": ["REFNO_UOB", "STATUS_POD", "RECEIVED/REASON", "CODING", "AWB", "REASON RETURN", "TGL_RECEIVED"],
    "YOUNG_LIVING_GROUP": ["AWB", "NOREF", "AMOUNT", "WEIGHT", "PROVINSI", "ETD", "TGL_RECEIVED", "TGL_ENTRY", "STATUS_POD"],
    "FIX_CONTACT_NOTELP_GROUP": ["CONTACT", "NOTELP"],
    "AGING_CARRER_GROUP": ["TGL_ENTRY", "1ST_ATTEMPT_DATE", "TGL_RECEIVED", "ETD", "STATUS_POD"]
}

# daftar kolom yang dibutuhkan (preserve)
REQUIRED_COLS = [
    "AWB","ID_ACCOUNT","SHIPPER_NAME","TGL_ENTRY","CONSIGNEE_NAME","ADDR1","ADDR2","ADDR3", "CONTACT","NOTELP","NOREF","ORIGIN","DEST","SERVICE","QTY","WEIGHT","GOODS_DESCR", "INSURANCE_ID","GOODS_VALUE","INSURANCE_VALUE(+)","AMOUNT","INTRUCTION","NOTICE", "HOLD_REASON","RECEIVING","RECEIVING_DATE","OUTBOUND_MANIFEST","OUTBOUND_MANIFEST_DATE",
    "INBOUND_MANIFEST","USER_IM","INBOUND_MANIFEST_DATE","MANIFEST_TRANSIT_AGEN","DATE_TRANSIT", "HVO_NO","HVO_DATE","HVO_HUB","HVO_HUB_NAME","HVO_HUB_DESTINATION","HVO_HUB_DESTINATION_NAME", "HVI_NO","HVI_DATE","RUNSHEET_NO","DATE_RUNSHEET","RUNSHEET_COURIER_ID","RUNSHEET_COURIER_NAME", "CODING","STATUS_POD","TGL_RECEIVED","STATUS_LATITUDE","STATUS_LONGITUDE","AGING","ETD","SLA", "CARRER","RECEIVED/REASON","TGL_UPDATE_STATUS_POD","WUS_OUTGOING_CODE","WUS_REMARKS","WUS_DATE", "INVOICED","AWB_CANCEL","COD_FLAG","BILNOTE_FLAG","BILNOTE_AMOUNT","REFNO_UOB","SCO_NO",
    "WO/DO/PO","NO_INVOICE","PAYMENT_TYPE","DATE_1ST_ATTEMPT","RESULT_1ST_ATTEMPT","LATLONG_1ST_ATTEMPT",     "DATE_2ND_ATTEMPT","RESULT_2ND_ATTEMPT","LATLONG_2ND_ATTEMPT","DATE_LAST_ATTEMPT","RESULT_LAST_ATTEMPT",     "LATLONG_LAST_ATTEMPT","PRA_RUNSHEET_NO","PRA_RUNSHEET_NAME","PRA_RUNSHEET_DATE","CS3_DATE", "CONNOTE_RETURN_RT","DATE_CONNOTE_RETURN_RT","CONNOTE_RETURN_RF","DATE_CONNOTE_RETURN_RF","USER_CONNOTE", "USER_ZONE_CONNOTE","CONFIRM_SHIPMENT_UNDEL","TRANSIT_MANIFEST","TRANSIT_MANIFEST_DATE","TRANSIT_MANIFEST_USER", "IREG_MANIFEST","IREG_CODE","IREG_DATE","URL_TTD","URL_FOTO","USER_OM","USER_RECEIVING","AGING_ONGOING", "CLAIM_NO","CLAIM_DOC_NO","CLAIM_DATE","NO_CNOTE_FW","ORIGIN_FW","DEST_FW","CODING_STATUS_FW","DESC_STATUS_FW", "HBG_NO","HBG_DATE","1ST_HVO_NO","1ST_HVO_DATE","1ST_HVO_USER","LAST_HVO_NO","LAST_HVO_DATE","LAST_HVO_USER", "MANIFEST_TRANSIT_SUBAGEN_NO", "MANIFEST_TRANSIT_SUBAGEN_DATE","MANIFEST_INBOUND_SUBAGEN_NO","MANIFEST_INBOUND_SUBAGEN_DATE", "BAG_NO","LATEST_SM_NO","LATEST_SM_DATE","1ST_PREVIOUS_SM_NO","1ST_PREVIOUS_SM_DATE","2ND_PREVIOUS_SM_NO","2ND_PREVIOUS_SM_DATE", "1ST_TRANSIT_MANIFEST_NO","1ST_TRANSIT_MANIFEST_DATE","2ND_TRANSIT_MANIFEST_NO","2ND_TRANSIT_MANIFEST_DATE", "3RD_TRANSIT_MANIFEST_NO","3RD_TRANSIT_MANIFEST_DATE","LAST_TRANSIT_MANIFEST_NO","LAST_TRANSIT_MANIFEST_DATE", "MTI_USER","MTS_USER","HO_COURIER_NO","HO_COURIER_DATE","WAREHOUSE_DATE","OFFICE_DATE","IRREG_REMAKS","BPIK","ZONE_USER_ENTRI", "CORRECT_DESTINATION","CORRECT_SERVICE","CORRECT_AMOUNT","HACB_NO","HACB_DATE","HACB_USER","HBAG_NO","HBAG_DATE","HBAG_USER", "PICKUP_DATE","PICKUP_STATUS","PICKUP_COURIER_ID","1ST_RUNSHEET_DATE","1ST_RUNSHEET_COURIERID","URL_CHAT","SINGLE_LEG", "LAST_DATE_DO","NO_RCW","DATE_RCW","USER_RCW","DATE_LPR","NO_LPR","NO_RDO","DATE_RDO","NO_DO","PROJECT_KR","HO_OFFICE_NO", "HO_OFFICE_DATE","LATEST_SM_ORIGIN","LATEST_SM_DEST","1ST_PREVIOUS_SM_ORIGIN","1ST_PREVIOUS_SM_DEST","2ND_PREVIOUS_SM_ORIGIN", "2ND_PREVIOUS_SM_DEST","TGL_TARIK_REPORT","RESPONCIBILITY","STATUS_VERSI_CCC","STATUS_POD_UPDATE","1ST_ATTEMPT_DATE","AGING_1ST", "CAREER_1ST","AGING_POD","CARRER_POD","CODING_UNDEL","REASON RETURN","3 LC DEST","NAMA KAB/KOTA 2","REGIONAL","ZONA", "GROUPING_SHIPPER","CATEGORY","REFERENCE CUST CCC","PAYMENT_METHODE","CUST_INDUSTRY","BIG_GROUPING_CUST","PIC_NAME_NEW RELATION", "PIC SUPPORT DATA","UNIT","DEPT","DATE","PERIODE","PERIODE_WEEK","ORIGIN RT","DEST RT","3LC Last Status","Regional Last Status", "Zona Last Status","OTS by CT","PERIODE_OTS","CLOSING_OTS","CODING_RT","RECEIVED/REASON_RT","Return Date"
]

table_reference_path = TABLE_REFERENCE_PATH

# ------------------- SAVE DENGAN SHORT DATE -------------------
def save_with_shortdate(
    df: pd.DataFrame,
    output_path: str,
    date_columns: list[str] | None = None,
):
    date_columns = date_columns or []

    with pd.ExcelWriter(
        output_path,
        engine="xlsxwriter",
        datetime_format="mm/dd/yyyy",
        date_format="mm/dd/yyyy",
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Sheet1",
            index=False,
            startrow=1,
            header=False
        )

        workbook  = writer.book
        worksheet = writer.sheets["Sheet1"]

        # --- formats ---
        header_fmt = workbook.add_format({
            "bold": True,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "bg_color": "#DCE6F1",
        })

        date_fmt = workbook.add_format({
            "num_format": "mm/dd/yyyy"
        })

        # --- write header with style ---
        worksheet.write_row(0, 0, df.columns.tolist(), header_fmt)

        # --- date column format ---
        for col_idx, col_name in enumerate(df.columns):
            if col_name in date_columns:
                worksheet.set_column(col_idx, col_idx, None, date_fmt)

        # --- autosize ---
        for i, col in enumerate(df.columns):
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(col)
            )
            worksheet.set_column(i, i, max_len + 2)

def style_worksheet(ws):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="DCE6F1",
        end_color="DCE6F1",
        fill_type="solid",
    )

    # header only
    for cell in ws[1]:
        cell.border = thin_border
        cell.fill = header_fill

def _update_status_history(df: pd.DataFrame, saved_path: str):
    if "STATUS_POD" not in df.columns:
        return
    today_counts = df["STATUS_POD"].value_counts().to_dict()
    basename = os.path.basename(str(saved_path))
    history = load_history()
    history[basename] = today_counts
    save_history(history)

def _safe_save_dataframe_excel(
    df: pd.DataFrame,
    output_path: str,
    *,
    use_shortdate: bool,
    max_retries: int = 3,
    delay: int = 1,
):
    attempt = 0
    base, ext = os.path.splitext(output_path)
    final_path = output_path

    while True:
        try:
            if use_shortdate:
                save_with_shortdate(df, final_path)
            else:
                df.to_excel(final_path, index=False)
            return final_path
        except PermissionError:
            attempt += 1
            if attempt > max_retries:
                raise
            final_path = f"{base} ({attempt}){ext}"
            print(f"⚠️ {output_path} sedang dipakai, coba simpan sebagai {final_path}")
            time.sleep(delay)

def insert_insurance_id(cols):
    cols = list(cols)
    if "INSURANCE_ID" in cols:
        return cols
    if "GOODS_DESCR" in cols:
        idx = cols.index("GOODS_DESCR") + 1
        return cols[:idx] + ["INSURANCE_ID"] + cols[idx:]
    existing_std = [c for c in REQUIRED_COLS if c in cols]
    if existing_std:
        last_std = existing_std[-1]
        idx = cols.index(last_std) + 1
        return cols[:idx] + ["INSURANCE_ID"] + cols[idx:]
    return ["INSURANCE_ID"] + cols

def dedupe_columns_preserve_order(cols):
    unique_cols = []
    seen = set()
    duplicates = []

    for col in cols:
        key = str(col).strip().upper()
        if key in seen:
            duplicates.append(str(col))
            continue
        seen.add(key)
        unique_cols.append(col)

    return unique_cols, duplicates

def strip_urls_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """Remove URL protocol prefix so openpyxl won't treat them as hyperlinks."""
    df = df.copy()
    url_cols = ["URL_TTD", "URL_FOTO", "URL_CHAT"]
    for col in url_cols:
        if col in df.columns:
            # convert URL to plain text (remove protocol, or truncate)
            df[col] = df[col].fillna("").astype(str).str.replace(r'^https?://', '', regex=True)
    return df

def _resolve_col_case_insensitive(df: pd.DataFrame, col_name: str | None) -> str | None:
    """Return actual column name in df that matches `col_name` case-insensitively."""
    if df is None or not isinstance(df, pd.DataFrame) or not col_name:
        return None

    target = str(col_name).strip().upper()
    for col in df.columns:
        if str(col).strip().upper() == target:
            return col
    return None


def _normalize_awb_key(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.strip()
        .str.lstrip("'")
        .str.replace(r"\s+", "", regex=True)
        .str.upper()
    )


def _clean_manual_values(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in df.columns:
        df[col] = df[col].astype("string").str.strip()
    return df.replace({"": pd.NA, "nan": pd.NA, "NaN": pd.NA, "None": pd.NA, "<NA>": pd.NA})


def _apply_manual_overrides(
    df: pd.DataFrame,
    manuals_path: str,
    *,
    debug: bool = False,
    announce: bool = True,
    stage: str = "",
) -> pd.DataFrame:
    manuals_df = pd.read_csv(manuals_path, dtype=str)
    manuals_df.columns = [str(c).strip().upper() for c in manuals_df.columns]

    if "AWB" not in manuals_df.columns:
        print("⚠️ manuals.csv harus punya kolom 'AWB'")
        return df
    if "AWB" not in df.columns:
        print("⚠️ Data master tidak punya kolom 'AWB', skip manuals.csv override")
        return df

    manuals_df = _clean_manual_values(manuals_df)
    manuals_df["_awb_key"] = _normalize_awb_key(manuals_df["AWB"])
    manuals_df = manuals_df[manuals_df["_awb_key"].notna() & (manuals_df["_awb_key"] != "")]

    manuals_df_deduped = manuals_df.drop(columns=["AWB"]).groupby("_awb_key", as_index=False).agg(
        lambda x: x.dropna().iloc[0] if len(x.dropna()) > 0 else pd.NA
    )

    duplicate_count = len(manuals_df) - len(manuals_df_deduped)
    if announce and duplicate_count > 0:
        if debug:
            print(f"[DEBUG] Found {duplicate_count} duplicate AWB entries, deduplicated to {len(manuals_df_deduped)}")
        print(f"📋 Deduped {duplicate_count} duplicate AWB entries dari manuals.csv")

    manuals_cols = [c for c in manuals_df_deduped.columns if c != "_awb_key"]
    if not manuals_cols:
        return df

    override_cols = {col: f"__manual_{col}" for col in manuals_cols}
    manuals_override = manuals_df_deduped[["_awb_key"] + manuals_cols].rename(columns=override_cols)

    df = df.copy()
    df["_awb_key"] = _normalize_awb_key(df["AWB"])
    df = df.merge(manuals_override, on="_awb_key", how="left")
    df.drop(columns=["_awb_key"], inplace=True, errors="ignore")

    applied_counts = {}
    for col in manuals_cols:
        manual_col = override_cols[col]
        if manual_col not in df.columns:
            continue

        manual_values = df[manual_col]
        mask = manual_values.notna()
        if col not in df.columns:
            df[col] = pd.NA
        df.loc[mask, col] = manual_values.loc[mask]

        # STATUS_POD is later rebuilt from STATUS_POD_UPDATE by add_status_pod_c01().
        # Keep both in sync so manual corrections survive that transform too.
        if col == "STATUS_POD" and "STATUS_POD_UPDATE" in df.columns:
            df.loc[mask, "STATUS_POD_UPDATE"] = manual_values.loc[mask]

        if mask.any():
            applied_counts[col] = int(mask.sum())
        df.drop(columns=[manual_col], inplace=True)

    if debug and applied_counts:
        stage_label = f" ({stage})" if stage else ""
        print(f"[DEBUG] Manual overrides applied{stage_label}: {applied_counts}")
    if announce:
        print(f"🔄 Data berhasil di-override dari {manuals_path} ({len(manuals_df_deduped)} unique AWBs)")

    return df


_KEY_FORMAT_FALLBACK_COLS = ["GOODS_DESCR", "NOTICE", "INTRUCTION", "NOREF"]


# ------------------- ADVANCED FILTER ENGINE -------------------
def _eval_single_condition(df: pd.DataFrame, condition: list) -> pd.Series:
    """Evaluate a single [col, operator, value] condition and return a boolean mask.

    Returns a Series of True for every row if the column is missing,
    so absent columns never accidentally exclude data.
    """
    col_name, ftype, arg = condition
    if col_name not in df.columns:
        return pd.Series(True, index=df.index)

    col = df[col_name]

    if ftype == "starts_with":
        return col.astype(str).str.startswith(str(arg))
    elif ftype == "ends_with":
        return col.astype(str).str.endswith(str(arg))
    elif ftype == "contains":
        return col.astype(str).str.contains(str(arg), case=False, na=False)
    elif ftype == "regex":
        return col.astype(str).str.contains(str(arg), case=False, na=False, regex=True)
    elif ftype == "equals":
        return col == arg
    elif ftype == "iequals":  # case-insensitive equals
        return col.astype(str).str.upper() == str(arg).upper()
    elif ftype == "not_equals":
        return col != arg
    elif ftype == "gt":
        return pd.to_numeric(col, errors="coerce") > float(arg)
    elif ftype == "lt":
        return pd.to_numeric(col, errors="coerce") < float(arg)
    elif ftype == "gte":
        return pd.to_numeric(col, errors="coerce") >= float(arg)
    elif ftype == "lte":
        return pd.to_numeric(col, errors="coerce") <= float(arg)
    elif ftype == "in":
        return col.isin(arg)
    elif ftype == "not_in":
        return ~col.isin(arg)
    else:
        raise ValueError(f"Unknown filter operator: {ftype}")


def apply_filters(df: pd.DataFrame, condition) -> pd.Series:
    """Recursively evaluate a nested filter condition tree, returning a boolean mask.

    *condition* can be:
      - A leaf:  ["COL", "operator", value]
      - AND:     {"and": [condition, ...]}
      - OR:      {"or":  [condition, ...]}
      - NOT:     {"not": condition}

    Leaves are evaluated via ``_eval_single_condition``.
    Logical nodes combine child masks with &, |, or ~.
    """
    # Leaf condition — a plain list/tuple of [col, op, value]
    if isinstance(condition, (list, tuple)) and len(condition) == 3 and isinstance(condition[0], str):
        return _eval_single_condition(df, condition)

    if not isinstance(condition, dict):
        raise ValueError(f"Invalid filter condition (expected dict or [col, op, val]): {condition}")

    # Exactly one logical key per dict node
    keys = [k.lower() for k in condition]
    if len(keys) != 1:
        raise ValueError(f"Filter dict must have exactly one key (and/or/not), got: {keys}")

    key = keys[0]
    value = condition[list(condition.keys())[0]]

    if key == "and":
        # value must be a list of sub-conditions
        masks = [apply_filters(df, sub) for sub in value]
        result = masks[0]
        for m in masks[1:]:
            result = result & m
        return result

    elif key == "or":
        masks = [apply_filters(df, sub) for sub in value]
        result = masks[0]
        for m in masks[1:]:
            result = result | m
        return result

    elif key == "not":
        # value is a single condition (dict or leaf)
        return ~apply_filters(df, value)

    else:
        raise ValueError(f"Unknown logical operator in filter: {key}")

def _collect_filter_columns(condition) -> set[str]:
    if isinstance(condition, (list, tuple)) and len(condition) == 3 and isinstance(condition[0], str):
        return {condition[0].strip().upper()}

    if not isinstance(condition, dict):
        return set()

    cols = set()
    for value in condition.values():
        if isinstance(value, list):
            for item in value:
                cols.update(_collect_filter_columns(item))
        else:
            cols.update(_collect_filter_columns(value))
    return cols


def _build_key_patterns(key_formats: list[str]) -> list[re.Pattern]:
    r"""Convert key_format examples into regex patterns.

    E.g. 'GEN.25/03/26-06' → re.compile(r'GEN\.\d+/\d+/\d+-\d+', re.I)
    Consecutive digits are replaced with \d+, everything else is escaped.
    """
    patterns = []
    for fmt in key_formats:
        parts = re.split(r'(\d+)', fmt)
        regex_str = "".join(
            r"\d+" if part.isdigit() else re.escape(part)
            for part in parts if part
        )
        patterns.append(re.compile(regex_str, re.IGNORECASE))
    return patterns


def _fill_key_from_fallback(
    df: pd.DataFrame,
    key_col: str,
    key_formats: list[str],
    fallback_cols: list[str] | None = None,
) -> int:
    """For rows where *key_col* is empty, try to extract a matching value
    from *fallback_cols* using patterns derived from *key_formats*.

    Modifies *df* in-place and returns the number of rows filled.
    """
    if not key_formats or key_col not in df.columns:
        return 0

    fallback_cols = fallback_cols or _KEY_FORMAT_FALLBACK_COLS
    # resolve actual column names (case-insensitive)
    resolved_fb = []
    for fb in fallback_cols:
        actual = _resolve_col_case_insensitive(df, fb)
        if actual and actual != key_col:
            resolved_fb.append(actual)
    if not resolved_fb:
        return 0

    patterns = _build_key_patterns(key_formats)

    blank_mask = (
        df[key_col].isna()
        | df[key_col].astype(str).str.strip().str.strip("'").str.strip().isin(["", "NAN", "NONE"])
    )
    filled = 0

    for idx in df.index[blank_mask]:
        found = None
        for fb_col in resolved_fb:
            val = str(df.at[idx, fb_col]).strip()
            if not val or val.upper() in ("NAN", "NONE", ""):
                continue
            for pat in patterns:
                m = pat.search(val)
                if m:
                    found = m.group(0).upper()
                    break
            if found:
                break
        if found:
            df.at[idx, key_col] = found
            filled += 1

    return filled

# ------------------- HELPERS: DuckDB-based IO -------------------
def get_recent_month_paths(base_dir, bulan_ke_belakang=3, category=str):
    today = datetime.today()
    folders = []
    for i in range(bulan_ke_belakang):
        target_date = today - relativedelta(months=i)
        tahun = target_date.year
        bulan = target_date.month
        nama_bulan = INDO_MONTHS[bulan]
        nama_folder_bulan = f"{bulan:02d}. {nama_bulan} {tahun}"

        # legacy layout (existing)
        legacy = os.path.join(base_dir, str(tahun), nama_folder_bulan, "CATEGORY", category)
        folders.append(legacy)

        # hive/partitioned layout (tahun=..., bulan=..., category=...)
        partition = os.path.join(base_dir, f"tahun={tahun}", f"bulan={bulan:02d}", f"category={category}")
        folders.append(partition)

    # keep order: newest-month legacy then partition, etc.
    return folders

def _collect_parquet_files(folders, base_dir=None, category=None, debug=False):
    """
    Collect parquet files from explicit `folders` first (legacy & partition paths).
    For each explicit folder we look:
      - parquet files directly inside the folder
      - parquet files inside immediate subfolders (one-level deep), e.g. OPEN/CLOSE
    If none found, fallback to a limited recursive search under `base_dir`.
    """
    files = []

    for folder in folders:
        try:
            if not os.path.exists(folder):
                if debug:
                    print(f"[DEBUG] folder not found: {folder}")
                continue

            folder_count = 0
            # files directly in folder
            for entry in os.listdir(folder):
                path = os.path.join(folder, entry)
                if os.path.isfile(path) and entry.lower().endswith(".parquet"):
                    files.append(path)
                    folder_count += 1
                elif os.path.isdir(path):
                    # one-level deep: check files inside this subfolder
                    try:
                        for sub in os.listdir(path):
                            sub_path = os.path.join(path, sub)
                            if os.path.isfile(sub_path) and sub.lower().endswith(".parquet"):
                                files.append(sub_path)
                                folder_count += 1
                    except Exception:
                        # ignore unreadable subfolders
                        continue
            if debug:
                print(f"[DEBUG] scanned folder: {folder} -> found {folder_count} parquet file(s)")
        except Exception:
            continue

    # 2) fallback: recursive search under base_dir only if explicit search found nothing
    if not files and base_dir and category:
        if debug:
            print(f"[DEBUG] explicit search found nothing, attempting limited recursive scan under {base_dir} for category={category}")
        cat_low = category.lower()
        for root, _, filenames in os.walk(base_dir):
            root_low = root.replace("\\", "/").lower()
            # only consider folders that look like partition or named category to reduce noise
            if ("category=" in root_low and f"category={cat_low}" in root_low) or \
               (f"/category/{cat_low}" in root_low) or root_low.endswith(f"/{cat_low}"):
                for fname in filenames:
                    if fname.lower().endswith(".parquet"):
                        files.append(os.path.join(root, fname))

        # last-resort: any parent folder named exactly category
        if not files:
            if debug:
                print(f"[DEBUG] limited recursive scan found nothing; doing last-resort search for parent folder named '{category}'")
            for root, _, filenames in os.walk(base_dir):
                if os.path.basename(root).lower() == cat_low:
                    for fname in filenames:
                        if fname.lower().endswith(".parquet"):
                            files.append(os.path.join(root, fname))

    unique_files = sorted(set(files))
    if debug:
        print(f"[DEBUG] total parquet files collected: {len(unique_files)}")
    return unique_files

def get_duckdb_connection(debug: bool = False):
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
    duckdb_temp_dir = os.path.join(project_root, "data", "duckdb_temp")
    os.makedirs(duckdb_temp_dir, exist_ok=True)
    os.environ["TEMP"] = duckdb_temp_dir
    os.environ["TMP"] = duckdb_temp_dir
    os.environ["TMPDIR"] = duckdb_temp_dir
    duckdb_temp_dir_sql = duckdb_temp_dir.replace("'", "''")

    con = duckdb.connect()
    con.execute("SET threads=4;")
    con.execute("SET preserve_insertion_order=false;")
    con.execute("SET memory_limit='5GB';")
    con.execute("SET max_temp_directory_size='200GB';")
    con.execute(f"SET temp_directory='{duckdb_temp_dir_sql}';")
    con.execute("SET enable_progress_bar=false;")
    con.execute("SET enable_object_cache=true;")

    if debug:
        settings = con.execute(
            """
            SELECT
                current_setting('temp_directory') AS temp_directory,
                current_setting('memory_limit') AS memory_limit,
                current_setting('max_temp_directory_size') AS max_temp_directory_size
            """
        ).fetchone()
        print(
            "[DEBUG] DuckDB settings: "
            f"temp_directory={settings[0]}, "
            f"memory_limit={settings[1]}, "
            f"max_temp_directory_size={settings[2]}"
        )

    return con

def _quote_duckdb_identifier(name: str) -> str:
    return f'"{str(name).replace(chr(34), chr(34) * 2)}"'

def _get_parquet_columns(con, files: list[str]) -> list[str]:
    if not files:
        return []

    info = con.execute("DESCRIBE SELECT * FROM read_parquet($1) LIMIT 0", [files]).fetchall()
    return [row[0] for row in info]

def _resolve_duckdb_columns(con, files: list[str], requested_cols=None) -> list[str]:
    available_cols = _get_parquet_columns(con, files)
    if not requested_cols:
        return available_cols

    available_by_upper = {
        str(col).strip().upper(): col
        for col in available_cols
    }

    resolved = []
    seen = set()
    for col in requested_cols:
        key = str(col).strip().upper()
        actual = available_by_upper.get(key)
        if actual and key not in seen:
            resolved.append(actual)
            seen.add(key)

    return resolved or available_cols

def _load_parquet_with_duckdb(con, parquet_files: list[str], query: str, columns=None):
    files = ",".join([f"'{f}'" for f in parquet_files])
    selected_columns = _resolve_duckdb_columns(con, parquet_files, columns)
    select_sql = ", ".join(_quote_duckdb_identifier(col) for col in selected_columns)
    sql = f"""
        SELECT {select_sql}
        FROM read_parquet([{files}])
        WHERE {query}
    """
    return con.execute(sql).df()

def _load_filtered_parquet_with_duckdb(
    con,
    files,
    date_col="TGL_ENTRY",
    start_dt=None,
    end_dt=None,
    id_account=None,
    selected_statuses=None,
    columns=None,
    debug=False,
):
    if not files:
        return pd.DataFrame()

    try:
        where_clauses = []
        params = []

        # date range
        if start_dt is not None:
            where_clauses.append(f'"{date_col}" >= ?')
            params.append(pd.Timestamp(start_dt).to_pydatetime())
        if end_dt is not None:
            where_clauses.append(f'"{date_col}" < ?')
            params.append(pd.Timestamp(end_dt).to_pydatetime())

        # ID_ACCOUNT: dukung data dengan/tanpa apostrophe di awal
        if id_account:
            id_params = []
            for x in id_account:
                raw = str(x).strip()
                if not raw:
                    continue
                no_quote = raw.lstrip("'")
                with_quote = raw if raw.startswith("'") else f"'{raw}"
                id_params.extend([with_quote, no_quote])

            id_params = list(dict.fromkeys(id_params))
            placeholders = ", ".join(["?"] * len(id_params))
            where_clauses.append(f'trim(cast("ID_ACCOUNT" as varchar)) IN ({placeholders})')
            params.extend(id_params)

        # STATUS_POD (case-insensitive match)
        if selected_statuses:
            placeholders = ", ".join(["?"] * len(selected_statuses))
            where_clauses.append(f"upper(STATUS_POD) IN ({placeholders})")
            params.extend([str(x).upper() for x in selected_statuses])

        where_sql = ""
        if where_clauses:
            where_sql = "WHERE " + " AND ".join(where_clauses)

        selected_columns = _resolve_duckdb_columns(con, files, columns)
        select_sql = ", ".join(_quote_duckdb_identifier(col) for col in selected_columns)

        sql = f"""
            SELECT {select_sql}
            FROM read_parquet($1)
            {where_sql}
        """

        if debug:
            print("[DEBUG SQL]", sql)
            print(f"[DEBUG] DuckDB selected columns: {len(selected_columns)}")

        df = con.execute(sql, [files] + params).df()
        if df.empty:
            return pd.DataFrame()

        df.columns = [c.strip().upper() for c in df.columns]
        return df

    except duckdb.OutOfMemoryException:
        raise
    except Exception as e:
        print(f"❌ DuckDB filtered read failed: {e}")
        return pd.DataFrame()

def sanitize_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    df = strip_urls_from_df(df)

    # --- object columns ---
    obj_cols = df.select_dtypes(include="object").columns

    for col in obj_cols:
        df[col] = df[col].apply(
            lambda x: (
                ", ".join(map(str, x)) if isinstance(x, (list, tuple, set))
                else str(x) if isinstance(x, dict)
                else x
            )
        )

    df[obj_cols] = df[obj_cols].where(df[obj_cols].notna(), None)

    # --- datetime columns ---
    dt_cols = df.select_dtypes(include=["datetime64[ns]", "datetimetz"]).columns

    for col in dt_cols:
        df[col] = df[col].map(
            lambda x: x.to_pydatetime() if isinstance(x, pd.Timestamp) else x
        )

    # Normalize common representations of missing/empty values so Excel shows empty cells
    # Replace pandas NA, numpy nan, literal "<NA>", and numeric 0 with None
    try:
        df = df.replace({pd.NA: None, np.nan: None, "<NA>": None, 0: None})
    except Exception:
        # best-effort: ignore if replace fails for some dtypes
        pass

    return df

def get_last_n_years(n=2):
    current_year = datetime.today().year
    return [current_year - i for i in range(n)][::-1]

# ------------------- MAIN FUNCTION -------------------
def get_data_from_master_pq(base_dir, criteria=None, category=str, output_dir=None, reference_path=str, bypass_history=False, debug=True, manuals_path=None):
    """
    DuckDB-first, per-criteria filtered pipeline.

    - If `criteria` is None: behaves like previous full load (loads recent months).
    - If `criteria` is provided (dict or list): for each crit it:
        * computes period window (period='year' or jumlah_bulan),
        * finds parquet files for months in that window (legacy + partition paths),
        * reads only columns available in parquet and only rows matching ID_ACCOUNT
          and date range using DuckDB (_load_filtered_parquet_with_duckdb),
        * normalizes dates, runs transforms and saves results as before.
    """
    base_unc = r"\\?\UNC" + base_dir[1:] if base_dir.startswith(r"\\") else base_dir
    con = get_duckdb_connection(debug=debug)

    # helper: create month list between start_dt (inclusive) and end_dt (exclusive)
    def _months_between(start_dt, end_dt):
        months = []
        cur = start_dt.replace(day=1)
        while cur < end_dt:
            months.append((cur.year, cur.month))
            cur = (cur + relativedelta(months=1)).replace(day=1)
        return months

    # ensure criteria is list
    if isinstance(criteria, dict):
        criteria = [criteria]

    save_base = output_dir or base_unc
    os.makedirs(save_base, exist_ok=True)
    had_errors = False

    for crit in criteria:
        try:
            group_name = crit.get("group_name")
            raw_id_account = crit.get("id_account")
            selected_cols = crit.get("selected_cols")
            exclude_cols = crit.get("exclude_cols")
            split_by_id = crit.get("split_by_id", False)
            split_by_col_val = crit.get("split_by_col_val")
            save_as = crit.get("save_as")
            jumlah_bulan = crit.get("jumlah_bulan")
            selected_statuses = crit.get("selected_statuses")
            filter_cols = crit.get("filter_cols")
            filters = crit.get("filters")  # new grouped-logic filters
            move_to_file = crit.get("move_to_sheet_of_file")
            clean_receiver = crit.get("clean_receiver")
            rename_cols = crit.get("rename_cols")
            saved_as = save_as or group_name
            cust_ref = crit.get("cust_ref") or {}
            selected_ref_cols = cust_ref.get("selected_ref_cols", [])
            ref_sheet = cust_ref.get("ref_sheet", [])
            col_order = cust_ref.get("col_order")
            period = (crit.get("period") or "").strip().lower()
            split_by_month_param = crit.get("split_by_month")
            if split_by_month_param is None:
                split_by_month = period not in {"last_n_months", "last_months", "month_range"}
            elif isinstance(split_by_month_param, str):
                split_by_month = split_by_month_param.strip().lower() not in {"0", "false", "no", "n", "off"}
            else:
                split_by_month = bool(split_by_month_param)

            print(f"\n🔍 Processing: {saved_as} (group: {group_name})")

            # =========================
            # DATE COLUMN DEFINITIONS
            # =========================
            MASTER_DATE_COL = "TGL_ENTRY"                 # parquet / master df
            REF_DATE_COL = cust_ref.get("date_col")       # e.g. TANGGAL PICKUP
            ACTIVE_DATE_COL = MASTER_DATE_COL 

            # --- normalize ID_ACCOUNT (pertahankan bentuk apostrophe) ---
            id_account = None
            if raw_id_account:
                id_account = [
                    str(x).strip()
                    for x in raw_id_account
                    if str(x).strip()
                ]
                
            process_key = f"query_{saved_as}"
            if move_to_file:
                process_key = f"query_{saved_as}__sheet_of_{move_to_file}"
            if not bypass_history and is_processed_today(process_key):
                print(f"⏩ Skip {saved_as}, sudah diproses hari ini (use bypass_history=True to override).")
                continue

            # --- compute date range for this crit ---
            if period == "year":
                # Query the last 1 year (inclusive). Later we split outputs by year.
                years_list = get_last_n_years(1)
                start_dt = datetime(int(years_list[0]), 1, 1)
                end_dt = datetime(int(years_list[-1]) + 1, 1, 1)
            else:
                # fallback: jumlah_bulan or 3
                jm = int(jumlah_bulan) if jumlah_bulan else 3
                today = datetime.today()
                end_dt = (today + relativedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
                start_month = (
                    today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                    - relativedelta(months=jm - 1)
                )
                start_dt = start_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

            if debug:
                print(f"[DEBUG] crit={saved_as} period start={start_dt} end={end_dt}")

            # --- build month-folder candidates for the months in the range ---
            months = _months_between(start_dt, end_dt)
            target_folders = []
            for (yr, mo) in months:
                nama_bulan = INDO_MONTHS[mo]
                nama_folder_bulan = f"{mo:02d}. {nama_bulan} {yr}"
                legacy = os.path.join(base_unc, str(yr), nama_folder_bulan, "CATEGORY", category)
                partition = os.path.join(base_unc, f"tahun={yr}", f"bulan={mo:02d}", f"category={category}")
                target_folders.extend([legacy, partition])
                if debug:
                    print(f"[DEBUG] added target folders: {legacy} , {partition}")

            # collect parquet files for those folders
            parquet_files = _collect_parquet_files(target_folders, base_unc, category, debug=debug)
            if not parquet_files:
                # last attempt: try scanning base_unc with fallback inside _collect_parquet_files
                parquet_files = _collect_parquet_files([], base_unc, category, debug=debug)

            if not parquet_files:
                print(f"⚠️ Tidak ada parquet file ditemukan untuk grup {saved_as} (folders checked).")
                continue

            if debug:
                print(f"[DEBUG] parquet files found for crit {saved_as}: {len(parquet_files)}")

            # desired columns: selected_cols + date_col + BIG_GROUPING_CUST + joins/filters
            desired = {
                str(col).strip().upper()
                for col in ((selected_cols or []) + [MASTER_DATE_COL, "BIG_GROUPING_CUST", "AWB", "TGL_TARIK_REPORT", "STATUS_POD", "ID_ACCOUNT"])
                if str(col).strip()
            }

            # add cust_ref key if present
            if cust_ref:
                key_left, key_right = cust_ref.get("key_col", ["AWB", "AWB"])
                desired.add(key_right.strip().upper() if isinstance(key_right, str) else key_right)

            # also include all cols referenced by transform groups/functions (uppercased)
            transform_needed = set()
            for tg_name, tg in TRANSFORM_GROUPS.items():
                group_output_cols = {str(col).strip().upper() for col in tg.get("cols", [])}
                if selected_cols and desired.intersection(group_output_cols):
                    for c in TRANSFORM_GROUP_INPUT_COLS.get(tg_name, []):
                        transform_needed.add(c.strip().upper())
            if selected_cols:
                for name in TRANSFORM_FUNCS:
                    key = name.strip().upper()
                    if key in desired:
                        for c in TRANSFORM_INPUT_COLS.get(name, []):
                            transform_needed.add(c.strip().upper())
            desired.update(transform_needed)

            if filter_cols:
                for col_name, _, _ in filter_cols:
                    desired.add(str(col_name).strip().upper())
            if filters:
                desired.update(_collect_filter_columns(filters))
            if split_by_col_val:
                desired.add(str(split_by_col_val).strip().upper())
            if clean_receiver:
                desired.update(["RECEIVED/REASON", "RECEIVING"])

            df = _load_filtered_parquet_with_duckdb(
                con,
                parquet_files,
                date_col=MASTER_DATE_COL,
                start_dt=start_dt,
                end_dt=end_dt,
                id_account=id_account,
                selected_statuses=selected_statuses,
                columns=desired,
                debug=debug
            )
            used_full_read_fallback = False
            # If DuckDB filter fails or returns empty, try full read
            if df is None or df.empty:
                if debug:
                    print(f"[DEBUG] DuckDB filtered read returned empty; attempting full read fallback.")
                fallback_query = "1=1"
                if id_account:
                    raw_vals = [str(x).strip() for x in id_account if str(x).strip()]
                    expanded_vals = []
                    for v in raw_vals:
                        no_quote = v.lstrip("'")
                        with_quote = v if v.startswith("'") else f"'{v}"
                        expanded_vals.extend([with_quote, no_quote])
                    expanded_vals = list(dict.fromkeys(expanded_vals))
                    # Escape quote agar nilai seperti '80044400 jadi literal SQL yang valid.
                    id_literal = ", ".join([f"'{v.replace("'", "''")}'" for v in expanded_vals])
                    fallback_query = f'trim(cast("ID_ACCOUNT" as varchar)) IN ({id_literal})'

                df_load = _load_parquet_with_duckdb(con, parquet_files, fallback_query, columns=desired)
                if df_load.empty:
                    print(f"⚠️ DuckDB full read returned empty for {saved_as}")
                    continue
                df = df_load.copy()
                used_full_read_fallback = True

            # --- Normalize column names to uppercase immediately and enforce filters ---
            df.columns = [c.strip().upper() for c in df.columns]

            # ensure date_col variable matches uppercase column name
            MASTER_DATE_COL = MASTER_DATE_COL.strip().upper()

            # Apply ID_ACCOUNT filter in pandas (MANDATORY SAFETY NET)
            if id_account and "ID_ACCOUNT" in df.columns:
                id_set = set()
                for x in id_account:
                    s = str(x).strip()
                    if not s:
                        continue
                    id_set.add(s)
                    id_set.add(s.lstrip("'"))
                    id_set.add(s if s.startswith("'") else f"'{s}")

                df["ID_ACCOUNT_NORM"] = (
                    df["ID_ACCOUNT"]
                    .astype(str)
                    .str.strip()
                )
                df = df[df["ID_ACCOUNT_NORM"].isin(id_set)]
                df.drop(columns=["ID_ACCOUNT_NORM"], inplace=True)

                if debug:
                    print(f"[DEBUG] After ID_ACCOUNT filter, rows={len(df)}")

            # Enforce date window in pandas only for full-read fallback path.
            # For normal path, DuckDB has already filtered by date and re-filtering may
            # drop rows due to pandas parse/coerce differences.
            if used_full_read_fallback and MASTER_DATE_COL in df.columns and start_dt is not None and end_dt is not None:
                try:
                    df[MASTER_DATE_COL] = pd.to_datetime(df[MASTER_DATE_COL], errors="coerce")
                    start_ts = pd.Timestamp(start_dt)
                    end_ts = pd.Timestamp(end_dt)
                    df = df[(df[MASTER_DATE_COL] >= start_ts) & (df[MASTER_DATE_COL] < end_ts)]
                    if debug:
                        print(f"[DEBUG] After fallback pandas date window ({start_ts} — {end_ts}), rows={len(df)}")
                except Exception as e:
                    if debug:
                        print(f"[DEBUG] Failed to enforce date window in pandas: {e}")
            else:
                if debug and not used_full_read_fallback:
                    print(f"[DEBUG] Using DuckDB date-filtered rows directly, rows={len(df)}")

            # normalize column names, dates & regional
            df.columns = [c.strip().upper() for c in df.columns]
            # convert date column to datetime if present
            if MASTER_DATE_COL in df.columns:
                try:
                    df[MASTER_DATE_COL] = pd.to_datetime(df[MASTER_DATE_COL], errors="coerce")
                except Exception:
                    pass

            # deduplicate AWB (keep latest by TGL_TARIK_REPORT if available else by date_col)
            if "AWB" in df.columns:
                before = len(df)
                if "TGL_TARIK_REPORT" in df.columns:
                    try:
                        df["TGL_TARIK_REPORT"] = pd.to_datetime(df["TGL_TARIK_REPORT"], errors="coerce")
                        df = df.sort_values(by=["AWB", "TGL_TARIK_REPORT"], ascending=[True, True])
                    except Exception:
                        pass
                elif MASTER_DATE_COL in df.columns:
                    try:
                        df = df.sort_values(by=["AWB", MASTER_DATE_COL], ascending=[True, True])
                    except Exception:
                        pass
                df = df.drop_duplicates(subset=["AWB"], keep="last")
                after = len(df)
                if before != after:
                    print(f"🧹 Duplikat dihapus: {before - after} baris (tinggal {after})")

            # sort by date_col (oldest to newest) for consistent ordering
            if MASTER_DATE_COL in df.columns:
                try:
                    df[MASTER_DATE_COL] = pd.to_datetime(df[MASTER_DATE_COL], errors="coerce")
                    df = df.sort_values(by=MASTER_DATE_COL, ascending=True)
                    if debug:
                        print(f"[DEBUG] Sorted by {MASTER_DATE_COL} (oldest to newest), rows={len(df)}")
                except Exception as e:
                    if debug:
                        print(f"[DEBUG] Failed to sort by {MASTER_DATE_COL}: {e}")

            # apply optional clean_receiver
            if clean_receiver and "RECEIVED/REASON" in df.columns:
                df["RECEIVED/REASON"] = df.apply(clean_receiver_column, axis=1)

            # --- Replace values pakai manuals.csv (sebelum cust_ref merge) ---
            _manuals_path = crit.get("manuals_path") or manuals_path
            if crit.get("edit_file", {}).get("replace_values") and _manuals_path and os.path.exists(_manuals_path):
                try:
                    if debug:
                        print(f"[DEBUG] Loading manuals.csv from: {_manuals_path}")
                    df = _apply_manual_overrides(df, _manuals_path, debug=debug, announce=True, stage="pre-transform")
                except Exception as e:
                    print(f"⚠️ Gagal replace values dari manuals.csv: {e}")

            # If a customer reference merge is requested and customer ref requires extra columns
            if cust_ref:
                ref_path = cust_ref.get("ref_path")
                key_left_raw, key_right_raw = cust_ref.get("key_col", ["AWB", "AWB"])
                jumlah_bulan_ref = cust_ref.get("jumlah_bulan", 3)

                print(f"🔄 Membaca customer reference dari: {ref_path}")
                df_custref = load_cust_ref(ref_path, jumlah_bulan_ref, ref_sheet, MASTER_DATE_COL)
                df_custref = normalize_all_dates(df_custref, debug=False)

                key_left = _resolve_col_case_insensitive(df_custref, key_left_raw)
                key_right = _resolve_col_case_insensitive(df, key_right_raw)

                # Backward-compatibility: handle criteria key order that is accidentally swapped
                # Expected order is [key_in_custref, key_in_master_df].
                if not (key_left and key_right):
                    swapped_left = _resolve_col_case_insensitive(df_custref, key_right_raw)
                    swapped_right = _resolve_col_case_insensitive(df, key_left_raw)
                    if swapped_left and swapped_right:
                        key_left, key_right = swapped_left, swapped_right
                        print(
                            f"ℹ️ key_col untuk {group_name} terdeteksi terbalik, "
                            f"otomatis pakai cust_ref='{key_left}' dan df='{key_right}'"
                        )

                # --- key_format fallback: fill blank key_right from other cols ---
                key_formats = cust_ref.get("key_format", [])
                if key_formats and key_right and key_right in df.columns:
                    fb_cols = cust_ref.get("key_fallback_cols") or None
                    filled = _fill_key_from_fallback(df, key_right, key_formats, fb_cols)
                    if filled:
                        print(f"🔑 key_format fallback: {filled} baris {key_right} diisi dari kolom lain")

                for df_, key in [(df_custref, key_left), (df, key_right)]:
                    if key in df_.columns:
                        df_[key] = (
                            df_[key]
                            .astype(str)
                            .str.strip()
                            .str.replace(r"\s+", "", regex=True)
                            .str.upper()
                            .apply(lambda x: x if x.startswith("'") else f"'{x}")
                        )

                if not df_custref.empty:
                    if key_left and key_right:
                        # Drop overlapping columns from df_custref that also exist in df
                        # (except merge key and explicitly requested ref cols) to prevent
                        # the left join from shadowing master data columns like AWB.
                        overlap_cols = set(df_custref.columns) & set(df.columns) - {key_left}
                        if selected_ref_cols:
                            overlap_cols -= {c.strip().upper() for c in selected_ref_cols}
                        if overlap_cols:
                            df_custref = df_custref.drop(columns=list(overlap_cols))
                        df_merged = df_custref.merge(df, how="left", left_on=key_left, right_on=key_right, suffixes=("", "_M"))
                        df_merged = df_merged[[c for c in df_merged.columns if not c.endswith("_M")]]
                        df = df_merged
                        if selected_ref_cols:
                            selected_cols = list(dict.fromkeys(selected_ref_cols + (selected_cols or [])))
                    else:
                        print(
                            f"⚠️ Skip merge cust_ref untuk {group_name}: "
                            f"key tidak ditemukan (left={key_left_raw}, right={key_right_raw}). "
                            f"key yang tersedia di cust_ref: {df_custref.columns.tolist()}, "
                            f"key yang tersedia di df: {df.columns.tolist()}"
                        )

                resolved_ref_date_col = _resolve_col_case_insensitive(df, REF_DATE_COL)
                ACTIVE_DATE_COL = resolved_ref_date_col or MASTER_DATE_COL

            else:
                ACTIVE_DATE_COL = MASTER_DATE_COL

            # =========================
            # FINAL DATE NORMALIZATION
            # =========================
            if ACTIVE_DATE_COL in df.columns:
                df[ACTIVE_DATE_COL] = pd.to_datetime(df[ACTIVE_DATE_COL], errors="coerce")

            if df.empty:
                print(f"⚠️ Tidak ada data cocok untuk grup: {group_name}")
                continue

            # continue with existing transform + save logic using `df` as df_filtered
            df_filtered = df.copy()

            # transformasi kolom (selected_cols etc.)

            # Order from oldest to newest based on ACTIVE_DATE_COL if available, to ensure consistent ordering for transforms and output
            if ACTIVE_DATE_COL in df_filtered.columns:
                df_filtered = df_filtered.sort_values(by=ACTIVE_DATE_COL, ascending=True)

            if selected_cols:
                selected_cols = insert_insurance_id(selected_cols)
                selected_cols, duplicate_selected_cols = dedupe_columns_preserve_order(selected_cols)
                if duplicate_selected_cols and debug:
                    print(
                        f"[DEBUG] Duplicate selected_cols skipped for {saved_as}: "
                        f"{duplicate_selected_cols}"
                    )
                if "INSURANCE_ID" not in df_filtered.columns:
                    df_filtered["INSURANCE_ID"] = ""
                needed_cols = set(selected_cols)

                for tg_name, group in TRANSFORM_GROUPS.items():
                    if any(col in needed_cols for col in group["cols"]):
                        if df_filtered.empty:
                            continue
                        try:
                            df_filtered = group["func"](df_filtered, reference_path)
                        except Exception as e:
                            print(f"⚠️ Gagal menambahkan {tg_name}: {e}")

                for col in needed_cols:
                    if col in TRANSFORM_FUNCS:
                        if df_filtered.empty:
                            continue
                        try:
                            df_filtered = TRANSFORM_FUNCS[col](df_filtered, reference_path)
                        except Exception as e:
                            print(f"⚠️ Gagal menambahkan {col}: {e}")

                for col in selected_cols:
                    if col not in df_filtered.columns:
                        df_filtered[col] = ""

                if crit.get("edit_file", {}).get("replace_values") and _manuals_path and os.path.exists(_manuals_path):
                    try:
                        df_filtered = _apply_manual_overrides(
                            df_filtered,
                            _manuals_path,
                            debug=debug,
                            announce=False,
                            stage="post-transform",
                        )
                    except Exception as e:
                        print(f"⚠️ Gagal apply ulang manuals.csv setelah transform: {e}")

            # apply filter_cols rules (legacy format — implicit AND)
            if filter_cols:
                for col_name, ftype, arg in filter_cols:
                    mask = _eval_single_condition(df_filtered, [col_name, ftype, arg])
                    df_filtered = df_filtered[mask]

            # apply advanced grouped filters (new format — supports AND/OR/NOT)
            if filters:
                mask = apply_filters(df_filtered, filters)
                df_filtered = df_filtered[mask]

            # prepare saver helper
            def prepare_df_for_save(df_segment: pd.DataFrame) -> pd.DataFrame:
                df_out = df_segment.copy()

                if "NO" in df.columns:
                    df_out = add_no(df_out)

                if exclude_cols:
                    df_out.drop(columns=exclude_cols, errors="ignore", inplace=True)

                if selected_cols:
                    cols_to_keep = [col for col in selected_cols if col in df_out.columns]
                    df_out = df_out[cols_to_keep]

                    if rename_cols:
                        if len(rename_cols) != len(selected_cols):
                            print(f"⚠️ Panjang rename_cols ≠ selected_cols untuk grup {group_name}, skip full rename")
                        else:
                            rename_map = {old: new for old, new in zip(selected_cols, rename_cols)}
                            rename_map = {k: v for k, v in rename_map.items() if k in df_out.columns}
                            if rename_map:
                                df_out.rename(columns=rename_map, inplace=True)
                                print(f"🔄 Kolom di-rename untuk grup {group_name}")
                else:
                    if rename_cols:
                        print(f"⚠️ rename_cols disediakan tanpa selected_cols untuk grup {group_name} → skip rename")

                # URUTKAN KOLOM SESUAI col_order (OPTIONAL)
                if col_order:
                    ordered_cols = [c for c in col_order if c in df_out.columns]
                    remaining_cols = [c for c in df_out.columns if c not in ordered_cols]
                    # df_out = df_out[ordered_cols + remaining_cols]
                    df_out = df_out[[c for c in col_order if c in df_out.columns]]

                    if "NO" in df_out.columns:
                        df_out = add_no(df_out)

                return df_out

            edit_file = crit.get("edit_file", {})

            # helper: save a dataframe segment using existing branching rules
            def _save_output(df_segment: pd.DataFrame, year_suffix: str | None = None):
                if df_segment is None or df_segment.empty:
                    return

                local_save_as = f"{save_as or group_name}"
                if year_suffix:
                    local_save_as = f"{local_save_as} - {year_suffix}"

                local_process_key = process_key + (f"_{year_suffix}" if year_suffix else "")

                if split_by_id and "ID_ACCOUNT" in df_segment.columns:
                    for id_acc, df_id in df_segment.groupby("ID_ACCOUNT"):
                        clean_id_acc = str(id_acc).lstrip("'")
                        if split_by_month and jumlah_bulan and ACTIVE_DATE_COL in df_id.columns:
                            df_id = df_id.copy()
                            df_id["BULAN_STR"] = df_id[ACTIVE_DATE_COL].dt.strftime("%B").str.upper()
                            for bulan_eng, df_bln in df_id.groupby("BULAN_STR"):
                                nama_bulan = EN_TO_ID_MONTH.get(bulan_eng, bulan_eng.title())
                                fname = f"{local_save_as} - {clean_id_acc} - {nama_bulan}.xlsx"
                                output_path = os.path.join(save_base, fname)
                                df_to_save = prepare_df_for_save(df_bln.drop(columns=["BULAN_STR"], errors="ignore"))
                                df_to_save = sanitize_df_for_excel(df_to_save)
                                output_path = _safe_save_dataframe_excel(
                                    df_to_save,
                                    output_path,
                                    use_shortdate=not bool(edit_file),
                                )
                                _update_status_history(df_bln if edit_file else df_to_save, output_path)
                                print(f"✅ Disimpan: {output_path}")
                                mark_processed_today(local_process_key)
                        else:
                            fname = f"{local_save_as} - {clean_id_acc}.xlsx"
                            output_path = os.path.join(save_base, fname)
                            df_to_save = prepare_df_for_save(df_id)
                            df_to_save = sanitize_df_for_excel(df_to_save)
                            output_path = _safe_save_dataframe_excel(
                                df_to_save,
                                output_path,
                                use_shortdate=not bool(edit_file),
                            )
                            _update_status_history(df_id if edit_file else df_to_save, output_path)
                            print(f"✅ Disimpan: {output_path}")
                            mark_processed_today(local_process_key)
                    return

                if split_by_col_val and split_by_col_val in df_segment.columns:
                    for val, df_val in df_segment.groupby(split_by_col_val):
                        clean_val = str(val).replace("/", "-").replace("\\", "-").strip()
                        if split_by_month and jumlah_bulan and ACTIVE_DATE_COL in df_val.columns:
                            df_val = df_val.copy()
                            df_val["BULAN_STR"] = df_val[ACTIVE_DATE_COL].dt.strftime("%B").str.upper()
                            for bulan_eng, df_bln in df_val.groupby("BULAN_STR"):
                                nama_bulan = EN_TO_ID_MONTH.get(bulan_eng, bulan_eng.title())
                                fname = f"{local_save_as} - {clean_val} - {nama_bulan}.xlsx"
                                output_path = os.path.join(save_base, fname)
                                df_to_save = prepare_df_for_save(df_bln.drop(columns=["BULAN_STR"], errors="ignore"))
                                df_to_save = sanitize_df_for_excel(df_to_save)
                                output_path = _safe_save_dataframe_excel(
                                    df_to_save,
                                    output_path,
                                    use_shortdate=not bool(edit_file),
                                )
                                _update_status_history(df_bln if edit_file else df_to_save, output_path)
                                print(f"✅ Disimpan: {output_path}")
                                mark_processed_today(local_process_key)
                        else:
                            fname = f"{local_save_as} - {clean_val}.xlsx"
                            output_path = os.path.join(save_base, fname)
                            df_to_save = prepare_df_for_save(df_val)
                            df_to_save = sanitize_df_for_excel(df_to_save)
                            output_path = _safe_save_dataframe_excel(
                                df_to_save,
                                output_path,
                                use_shortdate=not bool(edit_file),
                            )
                            _update_status_history(df_val if edit_file else df_to_save, output_path)
                            print(f"✅ Disimpan: {output_path}")
                            mark_processed_today(local_process_key)
                    return

                if move_to_file:
                    target_file = os.path.join(save_base, f"{move_to_file}.xlsx")
                    sheet_name = local_save_as
                    df_to_save = prepare_df_for_save(df_segment)
                    df_to_save = sanitize_df_for_excel(df_to_save)
                    if os.path.exists(target_file):
                        book = openpyxl.load_workbook(target_file)
                    else:
                        book = openpyxl.Workbook()
                        # Remove default empty "Sheet" created by openpyxl
                        if book.active and book.active.title == "Sheet":
                            book.remove(book.active)
                    if sheet_name in book.sheetnames:
                        book.remove(book[sheet_name])
                    ws = book.create_sheet(title=sheet_name)
                    for i, col in enumerate(df_to_save.columns, 1):
                        ws.cell(row=1, column=i, value=col)
                    for r_idx, row in enumerate(df_to_save.itertuples(index=False), 2):
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    style_worksheet(ws)
                    book.save(target_file)
                    print(f"📄 Data disimpan ke {target_file} (sheet: {sheet_name})")
                    mark_processed_today(local_process_key)
                    return

                if split_by_month and jumlah_bulan and ACTIVE_DATE_COL in df_segment.columns:
                    df_segment = df_segment.copy()
                    df_segment["BULAN_STR"] = df_segment[ACTIVE_DATE_COL].dt.strftime("%B").str.upper()
                    for bulan_eng, df_bln in df_segment.groupby("BULAN_STR"):
                        nama_bulan = EN_TO_ID_MONTH.get(bulan_eng, bulan_eng.title())
                        fname = f"{local_save_as} - {nama_bulan}.xlsx"
                        output_path = os.path.join(save_base, fname)
                        df_to_save = prepare_df_for_save(df_bln.drop(columns=["BULAN_STR"], errors="ignore"))
                        df_to_save = sanitize_df_for_excel(df_to_save)
                        output_path = _safe_save_dataframe_excel(
                            df_to_save,
                            output_path,
                            use_shortdate=not bool(edit_file),
                        )
                        _update_status_history(df_bln if edit_file else df_to_save, output_path)
                        print(f"✅ Disimpan: {output_path}")
                        mark_processed_today(local_process_key)
                    return

                # fallback: single file
                fname = f"{local_save_as}.xlsx"
                output_path = os.path.join(save_base, fname)
                df_to_save = prepare_df_for_save(df_segment)
                df_to_save = sanitize_df_for_excel(df_to_save)

                # Snapshot extra sheets before overwrite (from move_to_sheet_of_file)
                _extra_sheets = {}
                if os.path.exists(output_path):
                    try:
                        _tmpbook = openpyxl.load_workbook(output_path)
                        for _sn in _tmpbook.sheetnames:
                            if _sn == "Sheet1":
                                continue
                            _ws = _tmpbook[_sn]
                            _extra_sheets[_sn] = [[cell.value for cell in row] for row in _ws.iter_rows()]
                        _tmpbook.close()
                    except Exception:
                        pass

                output_path = _safe_save_dataframe_excel(
                    df_to_save,
                    output_path,
                    use_shortdate=not bool(edit_file),
                )

                # Re-add extra sheets that were preserved
                if _extra_sheets:
                    try:
                        _book = openpyxl.load_workbook(output_path)
                        for _sn, _rows in _extra_sheets.items():
                            if _sn in _book.sheetnames:
                                _book.remove(_book[_sn])
                            _ws2 = _book.create_sheet(title=_sn)
                            for _r_idx, _row_data in enumerate(_rows, 1):
                                for _c_idx, _val in enumerate(_row_data, 1):
                                    _ws2.cell(row=_r_idx, column=_c_idx, value=_val)
                            style_worksheet(_ws2)
                        _book.save(output_path)
                    except Exception:
                        pass

                _update_status_history(df_to_save, output_path)
                print(f"✅ Disimpan: {output_path}")
                mark_processed_today(local_process_key)

            # If period == 'year', split and save per year (last 2 years)
            if period == "year" and ACTIVE_DATE_COL in df_filtered.columns:
                years_list = get_last_n_years(1)
                for yr in years_list:
                    df_year = df_filtered[df_filtered[ACTIVE_DATE_COL].dt.year == int(yr)]
                    _save_output(df_year, year_suffix=str(yr))
            else:
                _save_output(df_filtered, year_suffix=None)

        except Exception as e:
            had_errors = True
            print(f"❌ Gagal proses {crit.get('group_name')}: {e}")
            continue

    con.close()
    return not had_errors

# keep existing get_data_from_rt (unchanged)
def get_data_from_rt(master_path: str, selected_cols: list, rt_master_path: str,
                     reference_path=table_reference_path, debug: bool=False) -> pd.DataFrame:
    if debug: print(f"[DEBUG] Load Master dari {master_path}")
    try:
        df_master = pd.read_excel(master_path, dtype=str)
    except Exception as e:
        print(f"⚠️ Gagal membaca master: {e}")
        return pd.DataFrame(columns=selected_cols)
    if "CONNOTE_RETURN_RT" not in df_master.columns:
        print("⚠️ Kolom CONNOTE_RETURN_RT tidak ada di Master!")
        return pd.DataFrame(columns=selected_cols)
    df_master = df_master[["CONNOTE_RETURN_RT"]].copy()
    if debug: print(f"[DEBUG] Load RT Master dari {rt_master_path}")
    try:
        df_rt = pd.read_excel(rt_master_path, dtype=str)
    except Exception as e:
        print(f"⚠️ Gagal membaca RT Master: {e}")
        return pd.DataFrame(columns=selected_cols)
    df_rt.columns = df_rt.columns.str.strip().str.upper()
    join_key = "AWB"
    if "AWB - COPY" in df_rt.columns:
        join_key = "AWB - COPY"
    df_joined = df_master.merge(
        df_rt,
        left_on="CONNOTE_RETURN_RT",
        right_on=join_key,
        how="inner"
    )
    if df_joined.empty:
        print("⚠️ Tidak ada AWB yang match di RT Master.")
        return pd.DataFrame(columns=selected_cols)
    try:
        df_ref = pd.read_excel(reference_path, dtype=str)
        df_ref.columns = df_ref.columns.str.strip().str.upper()
    except Exception as e:
        print(f"⚠️ Gagal membaca reference: {e}")
        df_ref = pd.DataFrame()
    df_out = df_joined.copy()
    for tg_name, group in TRANSFORM_GROUPS.items():
        if any(col in selected_cols for col in group["cols"]):
            if debug: print(f"Menambahkan kolom group {tg_name} ({group['cols']})..")
            try:
                df_out = group["func"](df_out, df_ref)
            except Exception as e:
                print(f"⚠️ Gagal menambahkan {tg_name}: {e}")
    for col in selected_cols:
        if col in TRANSFORM_FUNCS:
            if debug: print(f"Menambahkan kolom {col}..")
            try:
                df_out = TRANSFORM_FUNCS[col](df_out, df_ref)
            except Exception as e:
                print(f"⚠️ Gagal menambahkan {col}: {e}")
    for col in selected_cols:
        if col not in df_out.columns:
            df_out[col] = ""
    df_out = df_out[[col for col in selected_cols if col in df_out.columns]]
    return df_out
